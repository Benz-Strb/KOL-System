import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl, psycopg2

EXCEL_PATH = r'C:\Users\USER-SHD-046\Downloads\Copy of 2026 Dreame Plan&Summary Update.xlsx'
DB_URL = "postgresql://postgres.hdrweioqqqpslsjizkci:Shd2025%21ffofo@aws-1-ap-northeast-1.pooler.supabase.com:5432/postgres"

wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws = wb['KOL Performance 2026']
merge_map = {}
for mr in ws.merged_cells.ranges:
    tl = ws.cell(mr.min_row, mr.min_col)
    for row in range(mr.min_row, mr.max_row + 1):
        for col in range(mr.min_col, mr.max_col + 1):
            merge_map[(row, col)] = tl.value
def gv(r, c): return merge_map.get((r, c), ws.cell(r, c).value)
def norm(v):
    if v is None: return ''
    try: return str(round(float(v), 2)).rstrip('0').rstrip('.')
    except: return str(v).strip()

# DB rows 513-525 (context around Kit placements 515, 517, 519, 521)
conn = psycopg2.connect(DB_URL)
cur = conn.cursor()
cur.execute("""
    SELECT p.id, k.handle, plt.name, c.code, p.follower_at_time, p.final_price, p.payment_type
    FROM placements p JOIN kols k ON k.id=p.kol_id JOIN platforms plt ON plt.id=p.platform_id
    LEFT JOIN campaigns c ON c.id=p.campaign_id WHERE p.id BETWEEN 510 AND 530 ORDER BY p.id
""")
print("DB rows 510-530:")
for r in cur.fetchall():
    print(f"  id={r[0]:4d} handle={r[1]:<35} plt={r[2]:<12} camp={r[3]} fol={r[4]} price={r[5]} ptype={r[6]}")

cur.close()
conn.close()

print()
# Excel rows around those IDs (approximately rows 530-550 range)
print("Excel rows 530-560:")
for r in range(528, 562):
    kol = gv(r, 6)
    plat = str(gv(r, 3) or '')
    if kol:
        print(f"  row {r:4d}: campaign={norm(gv(r,1)):<6} platform={plat:<12} kol={str(kol):<35} fp={gv(r,11)}")

print()
# Search for base handle names in Excel
for base_handle in ['เบื้องหลังสมาร์ทโฮม', 'ติ่งไอที', 'ผมชอบแต่งบ้าน']:
    print(f"=== Search: {base_handle} ===")
    for r in range(3, ws.max_row+1):
        kol = gv(r, 6)
        if kol and base_handle in str(kol):
            print(f"  row={r:4d} campaign={norm(gv(r,1)):<6} platform={str(gv(r,3) or ''):<12} kol={str(kol):<35} fp={gv(r,11)}")
    print()
