"""
ตัดสินใจด้วย ±7 วันจากวัน campaign:
- ถ้า Excel corrected อยู่ใน ±7 วัน แต่ DB ไม่อยู่ → แก้ DB
- ถ้า DB อยู่ใน ±7 วัน แต่ Excel ไม่อยู่ → คง DB ไว้
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
import psycopg2
from datetime import datetime, date, timedelta
from collections import defaultdict

EXCEL_PATH = r'C:\Users\USER-SHD-046\Downloads\Copy of 2026 Dreame Plan&Summary Update.xlsx'
DB_URL = "postgresql://postgres.hdrweioqqqpslsjizkci:Shd2025%21ffofo@aws-1-ap-northeast-1.pooler.supabase.com:5432/postgres"

# Campaign code → (month, day) ของวัน campaign จริง
CAMPAIGN_DATE = {
    '1.1':(1,1),'1.15':(1,15),'1.25':(1,25),
    '2.2':(2,2),'2.15':(2,15),'2.25':(2,25),
    '3.3':(3,3),'3.15':(3,15),'3.25':(3,25),
    '4.4':(4,4),'4.15':(4,15),'4.25':(4,25),
    '5.5':(5,5),'5.15':(5,15),'5.25':(5,25),
    '6.6':(6,6),'6.15':(6,15),'6.25':(6,25),
    '7.7':(7,7),'7.15':(7,15),'7.25':(7,25),
    '8.8':(8,8),'8.15':(8,15),'8.25':(8,25),
    '9.9':(9,9),'9.15':(9,15),'9.25':(9,25),
    '10.10':(10,10),'10.15':(10,15),'10.25':(10,25),
    '11.11':(11,11),'11.15':(11,15),'11.25':(11,25),
    '12.12':(12,12),'12.15':(12,15),'12.25':(12,25),
}
WINDOW = 7  # ±7 วัน

def campaign_anchor(code, year=2026):
    md = CAMPAIGN_DATE.get(code)
    if md is None: return None
    return date(year, md[0], md[1])

def within_window(d, anchor, days=WINDOW):
    if d is None or anchor is None: return False
    return abs((d - anchor).days) <= days

def normalize_campaign(val):
    if val is None: return ''
    try:
        f = float(val)
        return f'{round(f,2):.2f}'.rstrip('0').rstrip('.')
    except: return str(val).strip()

def is_url(val):
    return bool(val) and str(val).strip().startswith('http')

def fix_date(d, campaign_code):
    if d is None: return None, 'none'
    if d.year < 2020 or d.year > 2030: return None, 'bad_year'
    anchor = campaign_anchor(campaign_code)
    if d.day > 12: return d, 'unambiguous'
    try:
        swapped = date(d.year, d.day, d.month)
    except ValueError:
        return d, 'swap_invalid'
    if anchor is None: return swapped, 'swapped_no_camp'
    orig_in  = within_window(d, anchor)
    swap_in  = within_window(swapped, anchor)
    if swap_in and not orig_in:  return swapped, 'swapped'
    if orig_in and not swap_in:  return d,       'original_ok'
    if swap_in and orig_in:      return swapped, 'both_ok_use_swap'
    return swapped, 'neither_ok_use_swap'

# Load Excel
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws = wb['KOL Performance 2026']
merge_map = {}
merge_top_left = {}
for mr in ws.merged_cells.ranges:
    tl_row, tl_col = mr.min_row, mr.min_col
    for row in range(mr.min_row, mr.max_row+1):
        for col in range(mr.min_col, mr.max_col+1):
            merge_map[(row,col)] = ws.cell(tl_row,tl_col).value
            merge_top_left[(row,col)] = (tl_row, tl_col)

def gv(r,c): return merge_map.get((r,c), ws.cell(r,c).value)

def get_pub_date_raw(row):
    ar, _ = merge_top_left.get((row,15),(row,15))
    raw = ws.cell(ar,15).value
    if is_url(raw): return None
    if isinstance(raw, datetime): return raw.date()
    if isinstance(raw, date): return raw
    for fmt in ('%d/%m/%Y','%d/%m/%y','%Y-%m-%d'):
        try:
            d = datetime.strptime(str(raw).strip(), fmt).date()
            if 2020 <= d.year <= 2030: return d
        except: pass
    return None

excel_rows = []
for r in range(3, ws.max_row+1):
    kol = gv(r,6); plat = gv(r,3)
    if not kol or not plat: continue
    ks, ps = str(kol).strip(), str(plat).strip()
    if ks==ps: continue
    if ks.lower().startswith(('total','รวม','sum')): continue
    camp = normalize_campaign(gv(r,1))
    excel_rows.append({'excel_row':r,'campaign':camp,'platform':ps,'handle':ks,
                       'raw_date':get_pub_date_raw(r)})

# DB
conn = psycopg2.connect(DB_URL)
cur = conn.cursor()
cur.execute("""
    SELECT p.id, k.handle, plt.name, c.code, p.publication_date, p.status
    FROM placements p JOIN kols k ON k.id=p.kol_id
    LEFT JOIN platforms plt ON plt.id=p.platform_id
    LEFT JOIN campaigns c ON c.id=p.campaign_id ORDER BY p.id
""")
db_by_3key = defaultdict(list)
for r in cur.fetchall():
    key = (r[1].lower(),(r[2] or '').lower(),r[3] or '')
    db_by_3key[key].append({'id':r[0],'handle':r[1],'platform':r[2] or '',
                             'campaign':r[3] or '','publication_date':r[4],'status':r[5]})

excel_used = set()
fixes = []
kept_db = []
no_anchor = []

for ep in excel_rows:
    k3 = (ep['handle'].lower(), ep['platform'].lower(), ep['campaign'])
    candidates = [d for d in db_by_3key.get(k3,[]) if d['id'] not in excel_used]
    if not candidates or ep['raw_date'] is None: continue
    dr = candidates[0]
    excel_used.add(dr['id'])
    corrected, reason = fix_date(ep['raw_date'], ep['campaign'])
    if corrected is None or dr['publication_date'] is None: continue
    if dr['publication_date'] == corrected: continue  # already ok

    anchor = campaign_anchor(ep['campaign'])
    if anchor is None:
        no_anchor.append(dr)
        continue

    db_in  = within_window(dr['publication_date'], anchor)
    cor_in = within_window(corrected, anchor)

    if cor_in and not db_in:
        fixes.append({**dr, 'new_date': corrected, 'reason': reason, 'anchor': anchor,
                      'db_diff': abs((dr['publication_date']-anchor).days),
                      'cor_diff': abs((corrected-anchor).days)})
    elif db_in:
        kept_db.append({**dr, 'corrected': corrected, 'anchor': anchor})
    else:
        no_anchor.append({**dr, 'corrected': corrected, 'anchor': anchor})

print(f"=== ผลด้วย ±{WINDOW} วัน จาก campaign date ===\n")
print(f"  แก้ (Excel ใกล้ campaign, DB ไม่ใกล้): {len(fixes)}")
print(f"  คง DB (DB ใกล้ campaign แล้ว):          {len(kept_db)}")
print(f"  ทั้งคู่ไม่ใกล้ / ไม่มี anchor:           {len(no_anchor)}")

print(f"\nFixes:")
# Group เพื่อแสดงให้กระชับ
sub = defaultdict(list)
for f in fixes:
    sub[(f['publication_date'], f['new_date'], f['campaign'])].append(f['id'])
for (db_d, new_d, camp), ids in sorted(sub.items()):
    anchor = campaign_anchor(camp)
    print(f"  campaign {camp} (anchor={anchor}): DB={db_d} → {new_d}  id={ids}")

if kept_db:
    print(f"\nDB kept (ถูกต้องแล้ว):")
    for m in kept_db:
        print(f"  id={m['id']:4d} [{m['handle'][:20]}/camp={m['campaign']}]: DB={m['publication_date']} (anchor={m['anchor']})")

if no_anchor:
    print(f"\nทั้งคู่ไม่ใกล้ campaign (ต้องดูเอง):")
    for m in no_anchor:
        anchor = m.get('anchor') or campaign_anchor(m.get('campaign',''))
        print(f"  id={m['id']:4d} [{m['handle'][:20]}/camp={m.get('campaign','')}]: DB={m['publication_date']}  corrected={m.get('corrected','-')}  anchor={anchor}")

# Generate SQL
sql_path = r'D:\internship\KOL_management\kol-system\scripts\pubdate_final_fixes.sql'
with open(sql_path, 'w', encoding='utf-8') as f:
    f.write(f"-- publication_date final fixes (±{WINDOW}-day window from campaign date)\n\n")
    for fix in fixes:
        f.write(f"UPDATE placements SET publication_date = '{fix['new_date']}' WHERE id = {fix['id']};\n")
print(f"\nSQL ({len(fixes)} statements) → {sql_path}")

cur.close()
conn.close()
