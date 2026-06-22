"""
Check specific DB rows (IDs 896-960) and ป้ายยาของแต่งบ้าน (444)
to understand why follower is NULL and what Excel has.
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
import psycopg2
import re
from decimal import Decimal, InvalidOperation

EXCEL_PATH = r'C:\Users\USER-SHD-046\Downloads\Copy of 2026 Dreame Plan&Summary Update.xlsx'
DB_URL = "postgresql://postgres.hdrweioqqqpslsjizkci:Shd2025%21ffofo@aws-1-ap-northeast-1.pooler.supabase.com:5432/postgres"

wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws = wb['KOL Performance 2026']
merge_map = {}
for mr in ws.merged_cells.ranges:
    tl = ws.cell(mr.min_row, mr.min_col)
    for row in range(mr.min_row, mr.max_row + 1):
        for col in range(mr.min_col, mr.max_col + 1):
            merge_map[(row, col)] = tl.value
def gv(row, col):
    return merge_map.get((row, col), ws.cell(row, col).value)

def normalize_campaign(val):
    if val is None: return ''
    try:
        f = float(val)
        s = f'{round(f, 2):.2f}'.rstrip('0').rstrip('.')
        return s
    except: return str(val).strip()

def parse_follower(val):
    if val is None: return None
    s = str(val).strip().upper().replace(',', '')
    if s in ('', 'NONE', '-', 'N/A', '0', '0.0'): return None
    try:
        m = re.match(r'^([\d.]+)\s*([KMB]?)$', s)
        if m:
            num = float(m.group(1))
            mult = {'K': 1000, 'M': 1_000_000, 'B': 1_000_000_000, '': 1}[m.group(2)]
            v = int(num * mult)
            return v if v > 0 else None
    except: pass
    return None

# Print raw Excel rows around where IDs 896-960 would have been imported
# (approximately rows 900-980 in Excel since there are header/section rows)
print("=== Excel rows 890-980 (looking for @-handle rows) ===")
for r in range(890, 985):
    kol = gv(r, 6)
    platform = gv(r, 3)
    campaign = normalize_campaign(gv(r, 1))
    follower = gv(r, 9)
    if kol:
        print(f"  Excel row {r:4d}: campaign={campaign:<6} platform={str(platform or ''):<12} "
              f"kol={str(kol):<35} follower={str(follower or '')}")

print()

# DB rows 896-960
conn = psycopg2.connect(DB_URL)
cur = conn.cursor()
cur.execute("""
    SELECT p.id, k.handle, plt.name, c.code, p.follower_at_time, p.final_price, p.payment_type
    FROM placements p
    JOIN kols k ON k.id = p.kol_id
    JOIN platforms plt ON plt.id = p.platform_id
    LEFT JOIN campaigns c ON c.id = p.campaign_id
    WHERE p.id BETWEEN 890 AND 965
    ORDER BY p.id
""")
print("=== DB rows 890-965 ===")
for row in cur.fetchall():
    print(f"  DB id={row[0]:4d} handle={row[1]:<35} platform={row[2]:<12} "
          f"campaign={row[3] or '':<6} follower={row[4]} price={row[5]}")

# Check id=444 specifically
print()
cur.execute("""
    SELECT p.id, k.handle, plt.name, c.code, p.follower_at_time, p.final_price, p.payment_type
    FROM placements p
    JOIN kols k ON k.id = p.kol_id
    JOIN platforms plt ON plt.id = p.platform_id
    LEFT JOIN campaigns c ON c.id = p.campaign_id
    WHERE p.id = 444
""")
row = cur.fetchone()
if row:
    print(f"DB id=444: handle={row[1]} platform={row[2]} campaign={row[3]} follower={row[4]}")

# Find all Excel rows for ป้ายยาของแต่งบ้าน
print("\nExcel rows for ป้ายยาของแต่งบ้าน:")
for r in range(3, ws.max_row + 1):
    kol = gv(r, 6)
    if kol and 'ป้ายยา' in str(kol):
        print(f"  Excel row {r}: campaign={normalize_campaign(gv(r,1))} "
              f"platform={gv(r,3)} kol={kol} follower={gv(r,9)} fp={gv(r,11)}")

# Find Excel rows for Crew Journey
print("\nExcel rows for Crew Journey:")
for r in range(3, ws.max_row + 1):
    kol = gv(r, 6)
    if kol and 'Crew' in str(kol):
        print(f"  Excel row {r}: campaign={normalize_campaign(gv(r,1))} "
              f"platform={gv(r,3)} kol={kol} follower={gv(r,9)} fp={gv(r,11)}")

# Find Excel rows for storybearcat
print("\nExcel rows for storybearcat:")
for r in range(3, ws.max_row + 1):
    kol = gv(r, 6)
    if kol and 'storybearcat' in str(kol).lower():
        print(f"  Excel row {r}: campaign={normalize_campaign(gv(r,1))} "
              f"platform={gv(r,3)} kol={kol} follower={gv(r,9)} fp={gv(r,11)}")

cur.close()
conn.close()
