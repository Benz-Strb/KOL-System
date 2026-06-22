"""
Read-only audit: find likely-duplicate KOL rows (same person, two `kols` rows)
and flag profile_url values that look like post links instead of profile links.

Does NOT modify the database. Reports findings for a human to decide on
(per project rule: never auto-merge ambiguous KOL records).
"""
import sys, io, re, unicodedata, difflib
from urllib.parse import urlparse, parse_qs
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import psycopg2

DB_URL = "postgresql://postgres.hdrweioqqqpslsjizkci:Shd2025%21ffofo@aws-1-ap-northeast-1.pooler.supabase.com:5432/postgres"

conn = psycopg2.connect(DB_URL)
cur = conn.cursor()
cur.execute("""
    SELECT k.id, k.handle, k.gen_name, k.profile_url, k.follower_count, k.avatar_url, p.name AS platform
    FROM kols k
    LEFT JOIN platforms p ON p.id = k.platform_id
    ORDER BY k.id
""")
rows = cur.fetchall()
cols = ['id', 'handle', 'gen_name', 'profile_url', 'follower_count', 'avatar_url', 'platform']
kols = [dict(zip(cols, r)) for r in rows]
print(f"Total KOL rows: {len(kols)}")

# ---- 1) normalize text (fold mathematical bold/italic unicode -> plain ascii, strip punctuation/space) ----
def normalize_text(s):
    if not s:
        return ''
    s = unicodedata.normalize('NFKD', s)
    s = re.sub(r'[​‌‍﻿]', '', s)  # zero-width junk
    s = re.sub(r'[^\w]', '', s, flags=re.UNICODE)
    return s.lower()

# ---- 2) normalize URL for exact-match comparison ----
def normalize_url(u):
    if not u:
        return None
    u = u.strip()
    p = urlparse(u if '://' in u else 'https://' + u)
    netloc = p.netloc.lower().replace('www.', '').replace('m.facebook.com', 'facebook.com')
    path = p.path.rstrip('/')
    # keep facebook profile.php?id=... as a distinguishing param
    qs = parse_qs(p.query)
    keep_q = ''
    if 'id' in qs and 'profile.php' in path:
        keep_q = f"?id={qs['id'][0]}"
    return f"{netloc}{path}{keep_q}"

# ---- 3) classify profile_url as profile-shaped vs post-shaped, per platform ----
POST_PATTERNS = {
    'facebook':  [r'/posts/', r'/videos/', r'/photos/', r'/watch/?\?', r'/reel/', r'permalink\.php', r'story_fbid', r'/video\.php'],
    'instagram': [r'/p/', r'/reel/', r'/tv/'],
    'tiktok':    [r'/video/\d+'],
    'youtube':   [r'/watch\?v=', r'/shorts/'],
    'lemon8':    [r'/p/'],
    'lamon8':    [r'/p/'],
}

def classify_url(platform, url):
    if not url:
        return 'none'
    plat = (platform or '').lower()
    patterns = POST_PATTERNS.get(plat, [])
    for pat in patterns:
        if re.search(pat, url, re.IGNORECASE):
            return 'post'
    return 'profile'

for k in kols:
    k['norm_url'] = normalize_url(k['profile_url'])
    k['url_kind'] = classify_url(k['platform'], k['profile_url'])
    k['norm_handle'] = normalize_text(k['handle'])
    k['norm_genname'] = normalize_text(k['gen_name'])

# ---- Finding A: exact same normalized profile_url across different KOL ids ----
SUFFIX_TOKENS = ['vdo', 'kit', 'photo', 'video']

def strip_format_suffix(norm):
    """Strip a trailing known content-format token (vdo/kit/photo/video) repeatedly."""
    changed = True
    while changed:
        changed = False
        for tok in SUFFIX_TOKENS:
            if norm.endswith(tok) and len(norm) > len(tok):
                norm = norm[: -len(tok)]
                changed = True
    return norm

def group_kind(group):
    """variant_pattern = all rows reduce to the same base after stripping vdo/kit/photo suffixes.
       needs_review = otherwise (genuinely different-looking names sharing one profile_url)."""
    bases = set()
    for k in group:
        base = strip_format_suffix(k['norm_handle']) or strip_format_suffix(k['norm_genname'])
        bases.add(base)
    return 'variant_pattern' if len(bases) <= 1 and all(bases) else 'needs_review'

by_url = {}
for k in kols:
    if k['url_kind'] != 'profile' or not k['norm_url']:
        continue
    by_url.setdefault(k['norm_url'], []).append(k)

groups_all = [(url, g) for url, g in by_url.items() if len(g) > 1]
groups_variant = [(u, g) for u, g in groups_all if group_kind(g) == 'variant_pattern']
groups_review = [(u, g) for u, g in groups_all if group_kind(g) == 'needs_review']

print(f"\n=== A1) Same profile_url, looks like intentional content-format split (main/VDO/Kit/Photo) — {len(groups_variant)} groups ===")
print("(probably NOT a bug — separate KOL row per content format on the same page. Listed for confirmation only.)")
for url, group in groups_variant:
    print(f"\n[{url}]")
    for k in group:
        print(f"  id={k['id']:<5} platform={k['platform']:<10} handle={k['handle']!r:30} gen_name={k['gen_name']!r:30} followers={k['follower_count']}")

print(f"\n=== A2) Same profile_url, NO simple format-suffix explanation — {len(groups_review)} groups — NEEDS HUMAN REVIEW ===")
for url, group in groups_review:
    print(f"\n[{url}]")
    for k in group:
        print(f"  id={k['id']:<5} platform={k['platform']:<10} handle={k['handle']!r:30} gen_name={k['gen_name']!r:30} followers={k['follower_count']}")
if not groups_review:
    print("(none found)")

# ---- Excel export of A2 for manual review ----
XLSX_PATH = 'kol_duplicates_needs_review.xlsx'
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'KOL duplicates'

HEADERS = ['group_no', 'shared_profile_url', 'kol_id', 'platform', 'handle', 'gen_name',
           'follower_count', 'actual_profile_url', 'keep_or_delete (กรอกเอง)', 'notes (กรอกเอง)']
ws.append(HEADERS)
header_fill = PatternFill('solid', fgColor='1F2937')
for cell in ws[1]:
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = header_fill
ws.freeze_panes = 'A2'

GROUP_FILLS = [PatternFill('solid', fgColor='F3F4F6'), PatternFill('solid', fgColor='FFFFFF')]
thin = Side(style='thin', color='D1D5DB')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

row_i = 2
for gi, (url, group) in enumerate(groups_review, start=1):
    group_sorted = sorted(group, key=lambda k: (k['follower_count'] is None, -(k['follower_count'] or 0)))
    fill = GROUP_FILLS[gi % 2]
    for k in group_sorted:
        ws.append([gi, url, k['id'], k['platform'], k['handle'], k['gen_name'],
                   k['follower_count'], k['profile_url'], '', ''])
        for cell in ws[row_i]:
            cell.fill = fill
            cell.border = border
        row_i += 1

dv = DataValidation(type='list', formula1='"เก็บ,ลบ,รวมเข้า id อื่น"', allow_blank=True)
ws.add_data_validation(dv)
dv.add(f'I2:I{row_i - 1}')

widths = [9, 30, 8, 10, 28, 26, 14, 45, 20, 30]
for col, w in zip('ABCDEFGHIJ', widths):
    ws.column_dimensions[col].width = w

wb.save(XLSX_PATH)
print(f"\nExcel exported: {XLSX_PATH} ({len(groups_review)} groups, {row_i - 2} rows)")

# ---- Finding B: same platform, fuzzy-similar handle/gen_name, DIFFERENT profile_url ----
# (pairs with the same profile_url are already covered by A1/A2 — only show the novel ones here)
print("\n=== B) Same platform, similar name but DIFFERENT profile_url — needs human judgement ===")
by_platform = {}
for k in kols:
    by_platform.setdefault(k['platform'], []).append(k)

THRESHOLD = 0.72
found_b = 0
seen_pairs = set()
for platform, group in by_platform.items():
    n = len(group)
    for i in range(n):
        for j in range(i + 1, n):
            a, b = group[i], group[j]
            if a['id'] == b['id']:
                continue
            # candidate text per record: handle and gen_name, compare best combination
            texts_a = [t for t in (a['norm_handle'], a['norm_genname']) if t]
            texts_b = [t for t in (b['norm_handle'], b['norm_genname']) if t]
            best = 0.0
            for ta in texts_a:
                for tb in texts_b:
                    if not ta or not tb:
                        continue
                    ratio = difflib.SequenceMatcher(None, ta, tb).ratio()
                    best = max(best, ratio)
            if best >= THRESHOLD:
                same_url = bool(a['norm_url'] and a['norm_url'] == b['norm_url'])
                if same_url:
                    continue  # already covered by A1/A2
                key = tuple(sorted((a['id'], b['id'])))
                if key in seen_pairs:
                    continue
                seen_pairs.add(key)
                found_b += 1
                print(f"\nsimilarity={best:.2f}  platform={platform}")
                for k in (a, b):
                    print(f"  id={k['id']:<5} handle={k['handle']!r:30} gen_name={k['gen_name']!r:30} "
                          f"profile_url={k['profile_url']!r:60} url_kind={k['url_kind']} followers={k['follower_count']}")
if found_b == 0:
    print("(none found)")

# ---- Finding C: profile_url that looks like a post link, not a profile link (separate data-quality issue) ----
print("\n=== C) profile_url looks like a POST link, not a profile link (by platform) ===")
post_like = [k for k in kols if k['url_kind'] == 'post']
print(f"Total: {len(post_like)} / {len(kols)} KOL rows")
from collections import Counter
print(Counter(k['platform'] for k in post_like))
print("\nSample (up to 15):")
for k in post_like[:15]:
    print(f"  id={k['id']:<5} platform={k['platform']:<10} handle={k['handle']!r:30} profile_url={k['profile_url']}")


# ---- Finding D: platform field doesn't match the domain in profile_url (likely wrong row got this URL) ----
print("\n=== D) platform field != domain of profile_url (data entry mismatch) ===")
DOMAIN_FOR_PLATFORM = {
    'facebook': 'facebook.com', 'instagram': 'instagram.com', 'tiktok': 'tiktok.com',
    'youtube': 'youtube.com', 'lemon8': 'lemon8', 'lamon8': 'lemon8',
}
found_d = 0
for k in kols:
    if not k['norm_url'] or not k['platform']:
        continue
    expected = DOMAIN_FOR_PLATFORM.get(k['platform'].lower())
    if expected and expected not in k['norm_url']:
        found_d += 1
        print(f"  id={k['id']:<5} platform={k['platform']:<10} handle={k['handle']!r:30} profile_url={k['profile_url']}")
if found_d == 0:
    print("(none found)")

cur.close()
conn.close()
