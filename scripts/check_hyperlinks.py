"""
ตรวจ hyperlink ใน column O (Post Link) ของ Excel
เปรียบเทียบกับ DB post_url
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
import psycopg2
from collections import defaultdict

EXCEL_PATH = r'C:\Users\USER-SHD-046\Downloads\Copy of 2026 Dreame Plan&Summary Update.xlsx'
DB_URL = "postgresql://postgres.hdrweioqqqpslsjizkci:Shd2025%21ffofo@aws-1-ap-northeast-1.pooler.supabase.com:5432/postgres"

def normalize_campaign(val):
    if val is None: return ''
    try:
        f = float(val)
        s = f'{round(f, 2):.2f}'.rstrip('0').rstrip('.')
        return s
    except:
        return str(val).strip()

def is_url(val):
    if not val: return False
    s = str(val).strip()
    return s.startswith('http://') or s.startswith('https://')

# Load Excel — ต้อง data_only=False เพื่อให้อ่าน hyperlink ได้
wb_hl = openpyxl.load_workbook(EXCEL_PATH, data_only=False)
ws_hl = wb_hl['KOL Performance 2026']

# Load อีกอันสำหรับ data_only=True (ค่า cell จริง + merge map)
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws = wb['KOL Performance 2026']

merge_map = {}
for mr in ws.merged_cells.ranges:
    tl = ws.cell(mr.min_row, mr.min_col)
    for row in range(mr.min_row, mr.max_row + 1):
        for col in range(mr.min_col, mr.max_col + 1):
            merge_map[(row, col)] = tl.value

def gv(row, col):
    return merge_map.get((row, col), ws.cell(row, col).value)

# ---- เก็บ post link จาก Excel (cell value + hyperlink) ----
excel_posts = []
for r in range(3, ws.max_row + 1):
    kol = gv(r, 6)
    platform = gv(r, 3)
    if not kol or not platform: continue
    kol_str = str(kol).strip()
    plat_str = str(platform).strip()
    if kol_str == plat_str: continue
    low = kol_str.lower()
    if low.startswith('total') or low.startswith('รวม') or low.startswith('sum'): continue

    cell_val = ws.cell(r, 15).value          # ค่าที่แสดงใน cell
    hl_cell  = ws_hl.cell(r, 15)
    hl_url   = hl_cell.hyperlink.target if hl_cell.hyperlink else None

    # เลือก URL ที่ดีที่สุด: hyperlink target > cell value (ถ้าเป็น URL)
    best_url = None
    if is_url(hl_url):
        best_url = str(hl_url).strip()
    elif is_url(cell_val):
        best_url = str(cell_val).strip()

    excel_posts.append({
        'excel_row': r,
        'campaign':  normalize_campaign(gv(r, 1)),
        'platform':  plat_str,
        'handle':    kol_str,
        'cell_val':  str(cell_val or '').strip(),
        'hl_url':    hl_url,
        'best_url':  best_url,
    })

has_hl      = [e for e in excel_posts if e['hl_url']]
has_url     = [e for e in excel_posts if e['best_url']]
hl_not_url  = [e for e in has_hl if not is_url(e['hl_url'])]

print(f"Total Excel data rows: {len(excel_posts)}")
print(f"  มี hyperlink:          {len(has_hl)}")
print(f"  มี URL (hl หรือ cell): {len(has_url)}")
print(f"  hyperlink ที่ไม่ใช่ http: {len(hl_not_url)}")

if hl_not_url:
    print("\nHyperlink ที่ไม่ใช่ URL (ตัวอย่าง 10 แถว):")
    for e in hl_not_url[:10]:
        print(f"  row={e['excel_row']:4d} [{e['handle'][:25]}/{e['platform']}] hl={e['hl_url']}  cell={e['cell_val']}")

# ---- เปรียบเทียบกับ DB ----
conn = psycopg2.connect(DB_URL)
cur = conn.cursor()
cur.execute("""
    SELECT p.id, k.handle, plt.name, c.code, p.post_url, p.status
    FROM placements p
    JOIN kols k ON k.id = p.kol_id
    LEFT JOIN platforms plt ON plt.id = p.platform_id
    LEFT JOIN campaigns c ON c.id = p.campaign_id
    ORDER BY p.id
""")
db_rows = {(r[1].lower(), (r[2] or '').lower(), r[3] or ''): []
           for r in []}  # init ว่าง
db_list = []
for r in cur.fetchall():
    db_list.append({'id': r[0], 'handle': r[1], 'platform': r[2] or '',
                    'campaign': r[3] or '', 'post_url': r[4], 'status': r[5]})

# Build lookup
db_by_3key = defaultdict(list)
for dr in db_list:
    k = (dr['handle'].lower(), dr['platform'].lower(), dr['campaign'])
    db_by_3key[k].append(dr)

excel_used = set()
fixes = []   # (db_id, handle, platform, campaign, new_url)
already_ok = 0
excel_null  = 0

for ep in excel_posts:
    k3 = (ep['handle'].lower(), ep['platform'].lower(), ep['campaign'])
    candidates = [d for d in db_by_3key.get(k3, []) if d['id'] not in excel_used]
    if not candidates:
        continue
    dr = candidates[0]
    excel_used.add(dr['id'])

    if ep['best_url'] is None:
        excel_null += 1
        continue

    db_url = (dr['post_url'] or '').strip()
    ex_url = ep['best_url'].strip()

    if db_url == ex_url:
        already_ok += 1
    else:
        fixes.append({
            'db_id':    dr['id'],
            'handle':   dr['handle'],
            'platform': dr['platform'],
            'campaign': dr['campaign'],
            'status':   dr['status'],
            'db_url':   db_url or 'NULL',
            'ex_url':   ex_url,
            'excel_row': ep['excel_row'],
        })

print(f"\n=== เปรียบเทียบ post_url ===")
print(f"  ตรงกันแล้ว:        {already_ok}")
print(f"  Excel ว่าง:         {excel_null}")
print(f"  ต้องแก้ไข (fix):   {len(fixes)}")

if fixes:
    # แยกตาม status
    planned_fixes = [f for f in fixes if f['status'] == 'planned']
    posted_fixes  = [f for f in fixes if f['status'] == 'posted']
    print(f"\n  ต้องแก้ (planned): {len(planned_fixes)}")
    print(f"  ต้องแก้ (posted):  {len(posted_fixes)}")

    print("\nDetail ทั้งหมด:")
    for f in fixes:
        print(f"  id={f['db_id']:4d} [{f['handle'][:25]}/{f['platform']}/camp={f['campaign']}] status={f['status']}")
        print(f"    DB:    {f['db_url'][:80]}")
        print(f"    Excel: {f['ex_url'][:80]}")

    # Generate SQL
    sql_lines = []
    for f in fixes:
        escaped = f['ex_url'].replace("'", "''")
        sql_lines.append(f"UPDATE placements SET post_url = '{escaped}' WHERE id = {f['db_id']};")

    print(f"\n=== SQL ({len(sql_lines)} statements) ===")
    for s in sql_lines:
        print(s)

    sql_path = r'D:\internship\KOL_management\kol-system\scripts\posturl_fixes.sql'
    with open(sql_path, 'w', encoding='utf-8') as f_out:
        f_out.write("-- post_url fixes from hyperlink audit\n\n")
        for s in sql_lines:
            f_out.write(s + "\n")
    print(f"\nSQL written to: {sql_path}")
else:
    print("\n✓ post_url ตรงกับ Excel ทุกแถวที่ match ได้")

cur.close()
conn.close()
