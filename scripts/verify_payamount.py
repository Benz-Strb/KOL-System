import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\USER-SHD-046\Downloads\Copy of 2026 Dreame Plan&Summary Update.xlsx', data_only=True)
ws = wb['KOL Performance 2026']
merge_map = {}
for mr in ws.merged_cells.ranges:
    tl = ws.cell(mr.min_row, mr.min_col)
    for row in range(mr.min_row, mr.max_row + 1):
        for col in range(mr.min_col, mr.max_col + 1):
            merge_map[(row, col)] = tl.value
def gv(r, c): return merge_map.get((r, c), ws.cell(r, c).value)
def norm(v):
    if v is None: return ''
    try: return str(round(float(v), 2)).rstrip('0').rstrip('.')
    except: return str(v).strip()

# Bank.sata TikTok campaign 2.2
print('=== Bank.sata / TikTok ===')
for r in range(3, ws.max_row+1):
    kol = gv(r, 6)
    plat = str(gv(r, 3) or '').lower()
    if kol and 'bank.sata' in str(kol).lower() and plat == 'tiktok':
        print(f'  row={r} campaign={norm(gv(r,1))} follower={gv(r,9)} fp={gv(r,11)} pay_amount={gv(r,12)}')

# แนนนี่รีวิว Facebook campaign 5.5
print()
print('=== แนนนี่รีวิว / Facebook ===')
for r in range(3, ws.max_row+1):
    kol = gv(r, 6)
    plat = str(gv(r, 3) or '').lower()
    if kol and 'แนนนี่รีวิว' in str(kol) and plat == 'facebook':
        print(f'  row={r} campaign={norm(gv(r,1))} follower={gv(r,9)} fp={gv(r,11)} pay_amount={gv(r,12)}')
