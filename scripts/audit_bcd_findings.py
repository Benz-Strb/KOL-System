"""
Read-only audit: re-run Finding B/C/D (from check_kol_duplicates.py) against the
CURRENT state of `kols` (after A1/A2 merges) and export each to its own Excel
review file for the team to fill in decisions.

Does NOT modify the database. Does NOT touch kol_duplicates_needs_review.xlsx /
kol_duplicates_A1_remaining_review.xlsx (those are finished A1/A2 evidence files).
"""
import sys, io, re, os, difflib, unicodedata
from urllib.parse import urlparse, parse_qs
from collections import Counter

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import psycopg2

env_path = os.path.join(os.path.dirname(__file__), '..', 'server', '.env')
direct_url = None
with open(env_path, encoding='utf-8') as f:
    for line in f:
        m = re.match(r'^DIRECT_URL="?(.*?)"?\s*$', line)
        if m:
            direct_url = m.group(1)
            break

conn = psycopg2.connect(direct_url)
cur = conn.cursor()
cur.execute("""
    SELECT k.id, k.handle, k.gen_name, k.profile_url, k.follower_count, p.name AS platform
    FROM kols k
    LEFT JOIN platforms p ON p.id = k.platform_id
    ORDER BY k.id
""")
rows = cur.fetchall()
cols = ['id', 'handle', 'gen_name', 'profile_url', 'follower_count', 'platform']
kols = [dict(zip(cols, r)) for r in rows]
print(f"Total KOL rows: {len(kols)}")


def normalize_text(s):
    if not s:
        return ''
    s = unicodedata.normalize('NFKD', s)
    s = re.sub(r'[​‌‍﻿]', '', s)
    s = re.sub(r'[^\w]', '', s, flags=re.UNICODE)
    return s.lower()


def normalize_url(u):
    if not u:
        return None
    u = u.strip()
    p = urlparse(u if '://' in u else 'https://' + u)
    netloc = p.netloc.lower().replace('www.', '').replace('m.facebook.com', 'facebook.com')
    path = p.path.rstrip('/')
    qs = parse_qs(p.query)
    keep_q = ''
    if 'id' in qs and 'profile.php' in path:
        keep_q = f"?id={qs['id'][0]}"
    return f"{netloc}{path}{keep_q}"


POST_PATTERNS = {
    'facebook':  [r'/posts/', r'/videos/', r'/photos/', r'/watch/?\?', r'/reel/', r'permalink\.php', r'story_fbid', r'/video\.php'],
    'instagram': [r'/p/', r'/reel/', r'/tv/'],
    'tiktok':    [r'/video/\d+'],
    'youtube':   [r'/watch\?v=', r'/shorts/'],
    'lemon8':    [r'/p/'],
    'lamon8':    [r'/p/'],
}


def classify_url(platform, url):
    if not url:
        return 'none'
    plat = (platform or '').lower()
    for pat in POST_PATTERNS.get(plat, []):
        if re.search(pat, url, re.IGNORECASE):
            return 'post'
    return 'profile'


for k in kols:
    k['norm_url'] = normalize_url(k['profile_url'])
    k['url_kind'] = classify_url(k['platform'], k['profile_url'])
    k['norm_handle'] = normalize_text(k['handle'])
    k['norm_genname'] = normalize_text(k['gen_name'])

HEADER_FILL = PatternFill('solid', fgColor='1F2937')
GROUP_FILLS = [PatternFill('solid', fgColor='F3F4F6'), PatternFill('solid', fgColor='FFFFFF')]
THIN = Side(style='thin', color='D1D5DB')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_sheet(ws, headers, col_widths, dv_col=None, dv_options=None, group_col_idx=0, last_row=1):
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = HEADER_FILL
    ws.freeze_panes = 'A2'
    for col, w in zip([chr(65 + i) for i in range(len(headers))], col_widths):
        ws.column_dimensions[col].width = w
    if dv_col and dv_options and last_row >= 2:
        dv = DataValidation(type='list', formula1=f'"{dv_options}"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f'{dv_col}2:{dv_col}{last_row}')


SCRIPTS_DIR = os.path.dirname(__file__)

# ---- Finding B: same platform, fuzzy-similar handle/gen_name, DIFFERENT profile_url ----
by_platform = {}
for k in kols:
    by_platform.setdefault(k['platform'], []).append(k)

THRESHOLD = 0.72
pairs_b = []
seen_pairs = set()
for platform, group in by_platform.items():
    n = len(group)
    for i in range(n):
        for j in range(i + 1, n):
            a, b = group[i], group[j]
            texts_a = [t for t in (a['norm_handle'], a['norm_genname']) if t]
            texts_b = [t for t in (b['norm_handle'], b['norm_genname']) if t]
            best = 0.0
            for ta in texts_a:
                for tb in texts_b:
                    best = max(best, difflib.SequenceMatcher(None, ta, tb).ratio())
            if best >= THRESHOLD:
                same_url = bool(a['norm_url'] and a['norm_url'] == b['norm_url'])
                if same_url:
                    continue
                key = tuple(sorted((a['id'], b['id'])))
                if key in seen_pairs:
                    continue
                seen_pairs.add(key)
                pairs_b.append((best, platform, a, b))

print(f"\nFinding B: {len(pairs_b)} pairs")

wb_b = openpyxl.Workbook()
ws_b = wb_b.active
ws_b.title = 'Finding B'
headers_b = ['pair_no', 'similarity', 'platform', 'kol_id', 'handle', 'gen_name',
             'follower_count', 'profile_url',
             'same_person_or_not (กรอกเอง)', 'merge_into_id (กรอกเอง ถ้าเลือกรวม)', 'notes (กรอกเอง)']
ws_b.append(headers_b)
row_i = 2
pairs_b.sort(key=lambda t: -t[0])
for pi, (sim, platform, a, b) in enumerate(pairs_b, start=1):
    fill = GROUP_FILLS[pi % 2]
    for k in (a, b):
        ws_b.append([pi, round(sim, 2), platform, k['id'], k['handle'], k['gen_name'],
                     k['follower_count'], k['profile_url'], '', '', ''])
        for cell in ws_b[row_i]:
            cell.fill = fill
            cell.border = BORDER
        row_i += 1
style_sheet(ws_b, headers_b, [9, 11, 10, 8, 28, 26, 14, 45, 22, 24, 30],
            dv_col='I', dv_options='เป็นคนเดียวกัน รวมเข้า id อื่น,เป็นคนละคน', last_row=row_i - 1)
path_b = os.path.join(SCRIPTS_DIR, 'kol_finding_B_review.xlsx')
wb_b.save(path_b)
print(f"Exported: {path_b} ({len(pairs_b)} pairs, {row_i - 2} rows)")

# ---- Finding C: profile_url looks like a post link, not a profile link ----
post_like = [k for k in kols if k['url_kind'] == 'post']
print(f"\nFinding C: {len(post_like)} rows")
print(Counter(k['platform'] for k in post_like))

wb_c = openpyxl.Workbook()
ws_c = wb_c.active
ws_c.title = 'Finding C'
headers_c = ['kol_id', 'platform', 'handle', 'gen_name', 'follower_count',
             'current_profile_url (เป็นลิงก์โพสต์)', 'correct_profile_url (กรอกเอง)', 'notes (กรอกเอง)']
ws_c.append(headers_c)
row_i = 2
for k in post_like:
    ws_c.append([k['id'], k['platform'], k['handle'], k['gen_name'], k['follower_count'],
                 k['profile_url'], '', ''])
    for cell in ws_c[row_i]:
        cell.border = BORDER
    row_i += 1
style_sheet(ws_c, headers_c, [8, 10, 28, 26, 14, 50, 50, 30])
path_c = os.path.join(SCRIPTS_DIR, 'kol_finding_C_review.xlsx')
wb_c.save(path_c)
print(f"Exported: {path_c} ({len(post_like)} rows)")

# ---- Finding D: platform field != domain of profile_url ----
DOMAIN_FOR_PLATFORM = {
    'facebook': 'facebook.com', 'instagram': 'instagram.com', 'tiktok': 'tiktok.com',
    'youtube': 'youtube.com', 'lemon8': 'lemon8', 'lamon8': 'lemon8',
}
mismatched = []
for k in kols:
    if not k['norm_url'] or not k['platform']:
        continue
    expected = DOMAIN_FOR_PLATFORM.get(k['platform'].lower())
    if expected and expected not in k['norm_url']:
        mismatched.append(k)
print(f"\nFinding D: {len(mismatched)} rows")

wb_d = openpyxl.Workbook()
ws_d = wb_d.active
ws_d.title = 'Finding D'
headers_d = ['kol_id', 'platform_in_db', 'handle', 'gen_name', 'follower_count',
             'profile_url', 'fix_action (กรอกเอง)', 'correct_value (กรอกเอง)', 'notes (กรอกเอง)']
ws_d.append(headers_d)
row_i = 2
for k in mismatched:
    ws_d.append([k['id'], k['platform'], k['handle'], k['gen_name'], k['follower_count'],
                 k['profile_url'], '', '', ''])
    for cell in ws_d[row_i]:
        cell.border = BORDER
    row_i += 1
style_sheet(ws_d, headers_d, [8, 14, 28, 26, 14, 50, 22, 30, 30],
            dv_col='G', dv_options='แก้ platform_in_db,แก้ profile_url,อื่นๆ ดู notes', last_row=row_i - 1)
path_d = os.path.join(SCRIPTS_DIR, 'kol_finding_D_review.xlsx')
wb_d.save(path_d)
print(f"Exported: {path_d} ({len(mismatched)} rows)")

cur.close()
conn.close()
