"""
Deep audit: match Excel rows to DB placements by sequence order,
then find specific mismatches in follower_at_time, final_price, target_pub_date.
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
import psycopg2
import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

EXCEL_PATH = r'C:\Users\USER-SHD-046\Downloads\Copy of 2026 Dreame Plan&Summary Update.xlsx'
DB_URL = "postgresql://postgres.hdrweioqqqpslsjizkci:Shd2025%21ffofo@aws-1-ap-northeast-1.pooler.supabase.com:5432/postgres"

# ---- Load Excel merge-aware ----
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws = wb['KOL Performance 2026']

merge_map = {}
for mr in ws.merged_cells.ranges:
    tl = ws.cell(mr.min_row, mr.min_col)
    for row in range(mr.min_row, mr.max_row + 1):
        for col in range(mr.min_col, mr.max_col + 1):
            merge_map[(row, col)] = tl.value

def gv(row, col):
    return merge_map.get((row, col), ws.cell(row, col).value)

def parse_follower(val):
    """Convert '106K', '1.2M', 106000, '2500.0' -> int"""
    if val is None:
        return None
    s = str(val).strip().upper().replace(',', '')
    if s in ('', 'NONE', '-', 'N/A'):
        return None
    try:
        m = re.match(r'^([\d.]+)\s*([KMB]?)$', s)
        if m:
            num = float(m.group(1))
            mult = {'K': 1000, 'M': 1_000_000, 'B': 1_000_000_000, '': 1}[m.group(2)]
            return int(num * mult)
    except:
        pass
    return None

def parse_price(val):
    """Convert to Decimal or None; treat 'Free','Barter',0 specially."""
    if val is None:
        return None
    s = str(val).strip().upper()
    if s in ('FREE', 'BARTER', '-', '', 'N/A', 'NONE'):
        return None
    try:
        d = Decimal(s.replace(',', ''))
        return d if d > 0 else None
    except InvalidOperation:
        return None

def parse_date(val):
    """Convert date-like value to date object."""
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val.date() if isinstance(val, datetime) else val
    s = str(val).strip()
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%m/%d/%Y', '%d-%m-%Y'):
        try:
            return datetime.strptime(s, fmt).date()
        except:
            pass
    return None

def infer_payment_type(final_price_raw):
    if final_price_raw is None:
        return None
    s = str(final_price_raw).strip().upper()
    if s == 'FREE':
        return 'free'
    if s == 'BARTER':
        return 'barter'
    return 'paid'

# Build Excel data rows (skip section headers)
excel_rows = []
for r in range(3, ws.max_row + 1):
    kol = gv(r, 6)
    platform = gv(r, 3)
    campaign = gv(r, 1)
    if not kol or not platform:
        continue
    kol_str = str(kol).strip()
    plat_str = str(platform).strip()
    # Section header: campaign cell spans multiple rows with no sub-fields
    if kol_str == plat_str:
        continue
    # Skip if looks like a total/summary row
    if str(kol_str).lower().startswith('total') or str(kol_str).lower().startswith('รวม'):
        continue
    fp_raw = gv(r, 11)
    excel_rows.append({
        'excel_row': r,
        'campaign': gv(r, 1),
        'pic': gv(r, 2),
        'platform': plat_str,
        'model': gv(r, 5),
        'kol_handle': kol_str,
        'gen_name': gv(r, 7),
        'content_cat': gv(r, 8),
        'follower_raw': gv(r, 9),
        'follower_int': parse_follower(gv(r, 9)),
        'final_price_raw': fp_raw,
        'final_price': parse_price(fp_raw),
        'payment_type': infer_payment_type(fp_raw),
        'pay_amount': parse_price(gv(r, 12)),
        'ads_cost': parse_price(gv(r, 13)),
        'target_pub_date': parse_date(gv(r, 14)),
        'post_link': gv(r, 15),
    })

print(f"Excel data rows: {len(excel_rows)}")

# ---- DB ----
conn = psycopg2.connect(DB_URL)
cur = conn.cursor()

cur.execute("""
    SELECT p.id, k.handle, plt.name as platform, c.code as campaign_code,
           p.follower_at_time, p.final_price, p.pay_amount, p.ads_cost,
           p.payment_type, p.placement_type, p.status,
           p.product_id, p.person_in_charge_id, p.person_in_charge,
           p.target_pub_date, p.post_url, p.campaign_id
    FROM placements p
    JOIN kols k ON k.id = p.kol_id
    JOIN platforms plt ON plt.id = p.platform_id
    LEFT JOIN campaigns c ON c.id = p.campaign_id
    ORDER BY p.id
""")
rows = cur.fetchall()
cols = ['id','handle','platform','campaign_code','follower_at_time','final_price',
        'pay_amount','ads_cost','payment_type','placement_type','status',
        'product_id','person_in_charge_id','person_in_charge',
        'target_pub_date','post_url','campaign_id']
db_data = [dict(zip(cols, r)) for r in rows]
print(f"DB placements: {len(db_data)}")

# ---- target_pub_date stats ----
has_tpd = sum(1 for r in db_data if r['target_pub_date'] is not None)
print(f"\ntarget_pub_date non-NULL in DB: {has_tpd}")
has_tpd_excel = sum(1 for r in excel_rows if r['target_pub_date'] is not None)
print(f"target_pub_date non-NULL in Excel: {has_tpd_excel}")

# ---- Match Excel to DB by sequential order ----
# Strategy: zip in order (both should be in same order from original import)
# Filter out rows that were deleted (IDs: 239, 245, 456, 961-964)
deleted_ids = {239, 245, 456, 961, 962, 963, 964}
# Note: DB has 1233 rows, Excel has 1237. The diff of 4 likely means
# some additional rows beyond the 7 deleted ones were added back,
# OR the count in CLAUDE.md was slightly off.

print(f"\nDB rows: {len(db_data)}, Excel rows: {len(excel_rows)}")
print(f"Diff: {len(excel_rows) - len(db_data)}")

# ---- Try sequential matching using (handle, platform) as anchor ----
# Build a map: (handle_lower, platform_lower) -> list of (db_idx, db_row)
from collections import defaultdict
db_by_key = defaultdict(list)
for i, r in enumerate(db_data):
    key = (r['handle'].lower(), r['platform'].lower())
    db_by_key[key].append((i, r))

excel_by_key = defaultdict(list)
for i, r in enumerate(excel_rows):
    key = (r['kol_handle'].lower(), r['platform'].lower())
    excel_by_key[key].append((i, r))

# Find Excel rows where target_pub_date is not null and try to match to DB
print("\n=== Sample: Excel has target_pub_date, check DB match ===")
sample_count = 0
for er in excel_rows[:50]:
    if er['target_pub_date'] and sample_count < 10:
        key = (er['kol_handle'].lower(), er['platform'].lower())
        db_matches = db_by_key.get(key, [])
        print(f"  Excel row {er['excel_row']:4d}: {er['kol_handle']:<25} {er['platform']:<12} "
              f"tpd={er['target_pub_date']} -> DB matches: {[r['id'] for _,r in db_matches][:5]}")
        sample_count += 1

# ---- Detailed: follower NULLs that CAN be filled ----
print("\n=== follower_at_time NULL in DB but Excel has value ===")
null_fol_fixable = []
for db_r in db_data:
    if db_r['follower_at_time'] is not None:
        continue
    key = (db_r['handle'].lower(), db_r['platform'].lower())
    excel_matches = excel_by_key.get(key, [])
    followers = [e['follower_int'] for _, e in excel_matches if e['follower_int'] is not None]
    if len(set(followers)) == 1:  # single unambiguous value
        null_fol_fixable.append({
            'db_id': db_r['id'],
            'handle': db_r['handle'],
            'platform': db_r['platform'],
            'follower_to_set': followers[0],
        })
    elif len(set(followers)) > 1:
        print(f"  AMBIGUOUS id={db_r['id']:4d} {db_r['handle']:<25} {db_r['platform']:<12} "
              f"values={list(set(followers))}")

print(f"\nUnambiguous follower fixes: {len(null_fol_fixable)}")
for r in null_fol_fixable:
    print(f"  id={r['db_id']:4d} {r['handle']:<25} {r['platform']:<12} -> {r['follower_to_set']}")

# ---- Detailed: final_price NULLs (paid) that CAN be filled ----
print("\n=== final_price NULL (paid) in DB but Excel has value ===")
null_fp_fixable = []
for db_r in db_data:
    if db_r['final_price'] is not None or db_r['payment_type'] != 'paid':
        continue
    key = (db_r['handle'].lower(), db_r['platform'].lower())
    excel_matches = excel_by_key.get(key, [])
    prices = [e['final_price'] for _, e in excel_matches if e['final_price'] is not None]
    if len(set(str(p) for p in prices)) == 1:
        null_fp_fixable.append({
            'db_id': db_r['id'],
            'handle': db_r['handle'],
            'platform': db_r['platform'],
            'price_to_set': prices[0],
        })
    elif prices:
        print(f"  AMBIGUOUS id={db_r['id']:4d} {db_r['handle']:<25} {db_r['platform']:<12} "
              f"values={list(set(str(p) for p in prices))}")

print(f"\nUnambiguous final_price fixes: {len(null_fp_fixable)}")
for r in null_fp_fixable:
    print(f"  id={r['db_id']:4d} {r['handle']:<25} {r['platform']:<12} -> {r['price_to_set']}")

cur.close()
conn.close()
