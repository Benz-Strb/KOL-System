"""
ตรวจ post_url ของ id 695-696 (Excel rows 720-721)
และ re-audit post_url ทั้งหมดด้วย merge-cell aware สำหรับ col O ด้วย
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
import psycopg2
from collections import defaultdict

EXCEL_PATH = r'C:\Users\USER-SHD-046\Downloads\Copy of 2026 Dreame Plan&Summary Update.xlsx'
DB_URL = "postgresql://postgres.hdrweioqqqpslsjizkci:Shd2025%21ffofo@aws-1-ap-northeast-1.pooler.supabase.com:5432/postgres"

def normalize_campaign(val):
    if val is None: return ''
    try:
        f = float(val)
        s = f'{round(f, 2):.2f}'.rstrip('0').rstrip('.')
        return s
    except:
        return str(val).strip()

def is_url(val):
    if not val: return False
    s = str(val).strip()
    return s.startswith('http://') or s.startswith('https://')

# Load ทั้งสองแบบ
wb_hl = openpyxl.load_workbook(EXCEL_PATH, data_only=False)  # สำหรับ hyperlink
wb    = openpyxl.load_workbook(EXCEL_PATH, data_only=True)   # สำหรับ cell values
ws_hl = wb_hl['KOL Performance 2026']
ws    = wb['KOL Performance 2026']

# Build merge map (ทุก column รวมถึง col O)
merge_map = {}         # (row, col) -> top-left value
merge_top_left = {}    # (row, col) -> (top_left_row, top_left_col)
for mr in ws.merged_cells.ranges:
    tl_row, tl_col = mr.min_row, mr.min_col
    tl_val = ws.cell(tl_row, tl_col).value
    for row in range(mr.min_row, mr.max_row + 1):
        for col in range(mr.min_col, mr.max_col + 1):
            merge_map[(row, col)] = tl_val
            merge_top_left[(row, col)] = (tl_row, tl_col)

def gv(row, col):
    return merge_map.get((row, col), ws.cell(row, col).value)

def get_hyperlink(row, col):
    """ดึง hyperlink URL จาก cell — ถ้าเป็น merged cell ให้ดึงจาก top-left"""
    actual_row, actual_col = merge_top_left.get((row, col), (row, col))
    cell = ws_hl.cell(actual_row, actual_col)
    if cell.hyperlink:
        return cell.hyperlink.target
    return None

def best_url_for(row, col):
    hl = get_hyperlink(row, col)
    if is_url(hl):
        return str(hl).strip()
    val = gv(row, col)
    if is_url(val):
        return str(val).strip()
    return None

# ---- ดู rows 718-722 (บริเวณ id 694-696) ----
print("=== Excel rows 718–722 (บริเวณ id 694–696) ===")
for r in range(718, 723):
    kol  = gv(r, 6)
    plat = gv(r, 3)
    camp = normalize_campaign(gv(r, 1))
    pic  = gv(r, 2)
    raw_cell = ws.cell(r, 15).value
    hl       = get_hyperlink(r, 15)
    best     = best_url_for(r, 15)
    print(f"  row={r}  kol={kol}  platform={plat}  campaign={camp}  PIC={pic}")
    print(f"    col O raw cell = {raw_cell}")
    print(f"    col O hyperlink= {hl}")
    print(f"    best_url       = {best}")
    print()

# ---- Re-audit post_url ทั้งหมดด้วย merge-aware ----
print("=== Re-audit post_url (merge-aware col O) ===")
excel_posts = []
for r in range(3, ws.max_row + 1):
    kol = gv(r, 6)
    platform = gv(r, 3)
    if not kol or not platform: continue
    kol_str  = str(kol).strip()
    plat_str = str(platform).strip()
    if kol_str == plat_str: continue
    low = kol_str.lower()
    if low.startswith('total') or low.startswith('รวม') or low.startswith('sum'): continue

    excel_posts.append({
        'excel_row': r,
        'campaign':  normalize_campaign(gv(r, 1)),
        'platform':  plat_str,
        'handle':    kol_str,
        'best_url':  best_url_for(r, 15),
    })

# Compare vs DB
conn = psycopg2.connect(DB_URL)
cur  = conn.cursor()
cur.execute("""
    SELECT p.id, k.handle, plt.name, c.code, p.post_url, p.status
    FROM placements p
    JOIN kols k ON k.id = p.kol_id
    LEFT JOIN platforms plt ON plt.id = p.platform_id
    LEFT JOIN campaigns c ON c.id = p.campaign_id
    ORDER BY p.id
""")
db_by_3key = defaultdict(list)
for r in cur.fetchall():
    key = (r[1].lower(), (r[2] or '').lower(), r[3] or '')
    db_by_3key[key].append({'id': r[0], 'handle': r[1], 'platform': r[2] or '',
                             'campaign': r[3] or '', 'post_url': r[4], 'status': r[5]})

excel_used = set()
fixes = []
for ep in excel_posts:
    k3 = (ep['handle'].lower(), ep['platform'].lower(), ep['campaign'])
    candidates = [d for d in db_by_3key.get(k3, []) if d['id'] not in excel_used]
    if not candidates or ep['best_url'] is None:
        continue
    dr = candidates[0]
    excel_used.add(dr['id'])

    db_url = (dr['post_url'] or '').strip()
    ex_url = ep['best_url'].strip()
    if db_url != ex_url:
        fixes.append({**dr, 'ex_url': ex_url, 'excel_row': ep['excel_row']})

print(f"ต้องแก้ไข: {len(fixes)} แถว")
for f in fixes:
    print(f"  id={f['id']:4d} [{f['handle'][:25]}/{f['platform']}/camp={f['campaign']}] status={f['status']}")
    print(f"    DB:    {(f['post_url'] or 'NULL')[:80]}")
    print(f"    Excel: {f['ex_url'][:80]}")

if fixes:
    sql_path = r'D:\internship\KOL_management\kol-system\scripts\posturl_merge_fixes.sql'
    with open(sql_path, 'w', encoding='utf-8') as f_out:
        f_out.write("-- post_url fixes (merge-cell aware)\n\n")
        for f in fixes:
            escaped = f['ex_url'].replace("'", "''")
            f_out.write(f"UPDATE placements SET post_url = '{escaped}' WHERE id = {f['id']};\n")
    print(f"\nSQL written to: {sql_path}")

cur.close()
conn.close()
