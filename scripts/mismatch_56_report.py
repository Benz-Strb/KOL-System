"""
สรุป 56 แถวที่ publication_date ยังต่างกันระหว่าง DB กับ Excel (หลัง DD/MM correction)
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
import psycopg2
from datetime import datetime, date
from collections import defaultdict

EXCEL_PATH = r'C:\Users\USER-SHD-046\Downloads\Copy of 2026 Dreame Plan&Summary Update.xlsx'
DB_URL = "postgresql://postgres.hdrweioqqqpslsjizkci:Shd2025%21ffofo@aws-1-ap-northeast-1.pooler.supabase.com:5432/postgres"

CAMPAIGN_MONTH = {
    '1.1':1,'1.15':1,'1.25':1,'2.2':2,'2.15':2,'2.25':2,
    '3.3':3,'3.15':3,'3.25':3,'4.4':4,'4.15':4,'4.25':4,
    '5.5':5,'5.15':5,'5.25':5,'6.6':6,'6.15':6,'6.25':6,
    '7.7':7,'7.15':7,'7.25':7,'8.8':8,'8.15':8,'8.25':8,
    '9.9':9,'9.15':9,'9.25':9,'10.10':10,'10.15':10,'10.25':10,
    '11.11':11,'11.15':11,'11.25':11,'12.12':12,'12.15':12,'12.25':12,
}

def normalize_campaign(val):
    if val is None: return ''
    try:
        f = float(val)
        return f'{round(f,2):.2f}'.rstrip('0').rstrip('.')
    except: return str(val).strip()

def is_url(val):
    return bool(val) and str(val).strip().startswith('http')

def fix_date(d, campaign_code):
    if d is None: return None, 'none'
    if d.year < 2020 or d.year > 2030: return None, 'bad_year'
    camp_month = CAMPAIGN_MONTH.get(campaign_code)
    if d.day > 12: return d, 'unambiguous'
    try:
        swapped = date(d.year, d.day, d.month)
    except ValueError:
        return d, 'swap_invalid'
    if camp_month is None: return swapped, 'swapped_no_camp'
    orig_ok = abs(d.month - camp_month) <= 1
    swap_ok = abs(swapped.month - camp_month) <= 1
    if swap_ok and not orig_ok: return swapped, 'swapped'
    if orig_ok and not swap_ok: return d, 'original_ok'
    if swap_ok and orig_ok:     return swapped, 'both_ok_use_swap'
    return swapped, 'neither_ok_use_swap'

# Load Excel
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws = wb['KOL Performance 2026']
merge_map = {}
merge_top_left = {}
for mr in ws.merged_cells.ranges:
    tl_row, tl_col = mr.min_row, mr.min_col
    for row in range(mr.min_row, mr.max_row+1):
        for col in range(mr.min_col, mr.max_col+1):
            merge_map[(row,col)] = ws.cell(tl_row,tl_col).value
            merge_top_left[(row,col)] = (tl_row, tl_col)

def gv(r,c): return merge_map.get((r,c), ws.cell(r,c).value)

def get_pub_date_raw(row):
    ar, _ = merge_top_left.get((row,15),(row,15))
    raw = ws.cell(ar,15).value
    if is_url(raw): return None
    if isinstance(raw, datetime): return raw.date()
    if isinstance(raw, date): return raw
    for fmt in ('%d/%m/%Y','%d/%m/%y','%Y-%m-%d'):
        try:
            d = datetime.strptime(str(raw).strip(), fmt).date()
            if 2020 <= d.year <= 2030: return d
        except: pass
    return None

excel_rows = []
for r in range(3, ws.max_row+1):
    kol = gv(r,6); plat = gv(r,3)
    if not kol or not plat: continue
    ks, ps = str(kol).strip(), str(plat).strip()
    if ks==ps: continue
    if ks.lower().startswith(('total','รวม','sum')): continue
    camp = normalize_campaign(gv(r,1))
    raw_date = get_pub_date_raw(r)
    excel_rows.append({'excel_row':r,'campaign':camp,'platform':ps,'handle':ks,'raw_date':raw_date})

conn = psycopg2.connect(DB_URL)
cur = conn.cursor()
cur.execute("""
    SELECT p.id, k.handle, plt.name, c.code, p.publication_date, p.status
    FROM placements p JOIN kols k ON k.id=p.kol_id
    LEFT JOIN platforms plt ON plt.id=p.platform_id
    LEFT JOIN campaigns c ON c.id=p.campaign_id ORDER BY p.id
""")
db_by_3key = defaultdict(list)
for r in cur.fetchall():
    key = (r[1].lower(),(r[2] or '').lower(),r[3] or '')
    db_by_3key[key].append({'id':r[0],'handle':r[1],'platform':r[2] or '',
                             'campaign':r[3] or '','publication_date':r[4],'status':r[5]})

excel_used = set()
mismatches = []
for ep in excel_rows:
    k3 = (ep['handle'].lower(), ep['platform'].lower(), ep['campaign'])
    candidates = [d for d in db_by_3key.get(k3,[]) if d['id'] not in excel_used]
    if not candidates or ep['raw_date'] is None: continue
    dr = candidates[0]
    excel_used.add(dr['id'])
    corrected, reason = fix_date(ep['raw_date'], ep['campaign'])
    if corrected is None: continue
    db_date = dr['publication_date']
    if db_date is not None and db_date != corrected:
        camp_month = CAMPAIGN_MONTH.get(ep['campaign'])
        mismatches.append({
            'id': dr['id'],
            'handle': dr['handle'],
            'platform': dr['platform'],
            'campaign': dr['campaign'],
            'camp_month': camp_month,
            'status': dr['status'],
            'db_date': db_date,
            'excel_raw': ep['raw_date'],
            'corrected': corrected,
            'reason': reason,
        })

# Group by reason
from collections import Counter
reason_count = Counter(m['reason'] for m in mismatches)
print(f"รวม {len(mismatches)} แถว\n")
print("จำแนกตาม reason:")
for r,c in reason_count.most_common():
    print(f"  {r}: {c} แถว")

# Print by reason group
for reason in ['both_ok_use_swap','original_ok','neither_ok_use_swap','swapped']:
    group = [m for m in mismatches if m['reason']==reason]
    if not group: continue

    if reason == 'both_ok_use_swap':
        print(f"\n{'='*60}")
        print(f"กลุ่ม: both_ok_use_swap ({len(group)} แถว)")
        print("→ วันกำกวม (day≤12) ทั้ง DB และ corrected อยู่ใน ±1 เดือนของ campaign")
        print("  script เลือก corrected (Thai intent) แต่ DB ต่างกัน")
        print("  ต้องเลือกว่าจะเชื่อ DB หรือ Excel\n")
        # Group by (db_date, corrected, campaign) เพื่อย่อ
        sub = defaultdict(list)
        for m in group:
            sub[(m['db_date'], m['corrected'], m['campaign'])].append(m['id'])
        for (db_d, cor_d, camp), ids in sorted(sub.items()):
            cm = CAMPAIGN_MONTH.get(camp,'?')
            print(f"  campaign {camp} (เดือน {cm}): DB={db_d}  Excel={cor_d}  → id={ids}")

    elif reason == 'original_ok':
        print(f"\n{'='*60}")
        print(f"กลุ่ม: original_ok ({len(group)} แถว)")
        print("→ DB month ตรงกับ campaign, แต่ corrected ไม่ตรง — DB น่าจะถูก\n")
        for m in group:
            print(f"  id={m['id']:4d} [{m['handle'][:25]}/camp={m['campaign']}]: DB={m['db_date']}  Excel_corrected={m['corrected']}")

    elif reason == 'neither_ok_use_swap':
        print(f"\n{'='*60}")
        print(f"กลุ่ม: neither_ok_use_swap ({len(group)} แถว)")
        print("→ ทั้ง DB และ corrected ไม่ตรง campaign month เลย — ต้องดูเอง\n")
        for m in group:
            cm = m['camp_month']
            print(f"  id={m['id']:4d} [{m['handle'][:25]}/camp={m['campaign']} เดือน {cm}]: DB={m['db_date']}  Excel_raw={m['excel_raw']}  corrected={m['corrected']}")

    elif reason == 'swapped':
        print(f"\n{'='*60}")
        print(f"กลุ่ม: swapped ({len(group)} แถว)")
        print("→ DB month ไม่ตรง campaign แต่ corrected ตรง — DB น่าจะผิด ควรแก้\n")
        sub = defaultdict(list)
        for m in group:
            sub[(m['db_date'], m['corrected'], m['campaign'])].append(m['id'])
        for (db_d, cor_d, camp), ids in sorted(sub.items()):
            cm = CAMPAIGN_MONTH.get(camp,'?')
            print(f"  campaign {camp} (เดือน {cm}): DB={db_d}  Excel_corrected={cor_d}  → id={ids}")

cur.close()
conn.close()
