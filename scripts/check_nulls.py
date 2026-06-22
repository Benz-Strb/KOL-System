"""
Audit NULL fields in placements vs Excel merge-aware values.
Checks: follower_at_time, final_price, pay_amount, product_id,
        person_in_charge_id, person_in_charge (text), target_pub_date
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
import psycopg2

EXCEL_PATH = r'C:\Users\USER-SHD-046\Downloads\Copy of 2026 Dreame Plan&Summary Update.xlsx'
DB_URL = "postgresql://postgres.hdrweioqqqpslsjizkci:Shd2025%21ffofo@aws-1-ap-northeast-1.pooler.supabase.com:5432/postgres"

# ---- Load Excel merge-aware ----
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws = wb['KOL Performance 2026']

merge_map = {}
for mr in ws.merged_cells.ranges:
    tl = ws.cell(mr.min_row, mr.min_col)
    for row in range(mr.min_row, mr.max_row + 1):
        for col in range(mr.min_col, mr.max_col + 1):
            merge_map[(row, col)] = tl.value

def gv(row, col):
    return merge_map.get((row, col), ws.cell(row, col).value)

# Col indices (1-based)
# A=1 Campaign, B=2 PIC, C=3 Platform, D=4 Cat(product), E=5 Model,
# F=6 KOL, G=7 GenName, H=8 ContentCat, I=9 Follower, J=10 Tier,
# K=11 FinalPrice, L=12 PayAmt, M=13 AdsCost, N=14 TargetPubDate, O=15 PostLink

excel_rows = []
for r in range(3, ws.max_row + 1):
    kol = gv(r, 6)
    platform = gv(r, 3)
    campaign = gv(r, 1)
    if not kol or not platform:
        continue
    kol_str = str(kol).strip()
    # Skip section headers (merged campaign label rows)
    if '/' in str(campaign) and not platform:
        continue
    follower_raw = gv(r, 9)
    follower_str = str(follower_raw).strip() if follower_raw is not None else None
    excel_rows.append({
        'excel_row': r,
        'campaign': gv(r, 1),
        'pic': gv(r, 2),
        'platform': str(platform).strip(),
        'model': gv(r, 5),
        'kol_handle': kol_str,
        'gen_name': gv(r, 7),
        'content_cat': gv(r, 8),
        'follower_raw': follower_str,
        'final_price': gv(r, 11),
        'pay_amount': gv(r, 12),
        'ads_cost': gv(r, 13),
        'target_pub_date': gv(r, 14),
        'post_link': gv(r, 15),
    })

print(f"Excel data rows: {len(excel_rows)}")

# ---- DB ----
conn = psycopg2.connect(DB_URL)
cur = conn.cursor()

# Get all placements with nullable fields
cur.execute("""
    SELECT p.id, k.handle, plt.name as platform,
           p.follower_at_time, p.final_price, p.pay_amount, p.ads_cost,
           p.payment_type, p.placement_type, p.status,
           p.product_id, p.person_in_charge_id, p.person_in_charge,
           p.target_pub_date, p.post_url, p.campaign_id
    FROM placements p
    JOIN kols k ON k.id = p.kol_id
    JOIN platforms plt ON plt.id = p.platform_id
    ORDER BY p.id
""")
rows = cur.fetchall()
cols = ['id','handle','platform','follower_at_time','final_price','pay_amount',
        'ads_cost','payment_type','placement_type','status',
        'product_id','person_in_charge_id','person_in_charge',
        'target_pub_date','post_url','campaign_id']
db_data = [dict(zip(cols, r)) for r in rows]

print(f"DB placements: {len(db_data)}")
print()

# ---- NULL summary ----
checks = [
    ('follower_at_time', lambda r: r['follower_at_time'] is None),
    ('final_price (all)', lambda r: r['final_price'] is None),
    ('final_price (paid only)', lambda r: r['final_price'] is None and r['payment_type'] == 'paid'),
    ('pay_amount', lambda r: r['pay_amount'] is None),
    ('product_id (online)', lambda r: r['product_id'] is None and r['placement_type'] == 'online'),
    ('person_in_charge_id', lambda r: r['person_in_charge_id'] is None),
    ('person_in_charge (text)', lambda r: r['person_in_charge'] is None),
    ('target_pub_date', lambda r: r['target_pub_date'] is None),
    ('campaign_id', lambda r: r['campaign_id'] is None),
]
print("NULL counts in DB placements:")
for label, fn in checks:
    cnt = sum(1 for r in db_data if fn(r))
    print(f"  {label}: {cnt}")

print()

# ---- Build handle->rows index for Excel matching ----
from collections import defaultdict
excel_by_handle = defaultdict(list)
for er in excel_rows:
    excel_by_handle[er['kol_handle'].lower()].append(er)

# ---- Check follower_at_time NULLs ----
null_fol = [r for r in db_data if r['follower_at_time'] is None]
print(f"\n=== follower_at_time NULL ({len(null_fol)} rows) ===")
for r in null_fol[:30]:
    handle = r['handle']
    matches = excel_by_handle.get(handle.lower(), [])
    excel_vals = list({e['follower_raw'] for e in matches if e['follower_raw']})
    print(f"  id={r['id']:4d} handle={handle:<25} excel_follower={excel_vals}")

# ---- Check person_in_charge NULL ----
null_pic = [r for r in db_data if r['person_in_charge_id'] is None]
print(f"\n=== person_in_charge_id NULL ({len(null_pic)} rows) ===")
for r in null_pic[:30]:
    handle = r['handle']
    matches = excel_by_handle.get(handle.lower(), [])
    excel_pics = list({e['pic'] for e in matches if e['pic']})
    print(f"  id={r['id']:4d} handle={handle:<25} excel_pic={excel_pics}")

# ---- Check final_price NULL (paid) ----
null_fp_paid = [r for r in db_data if r['final_price'] is None and r['payment_type'] == 'paid']
print(f"\n=== final_price NULL (paid, {len(null_fp_paid)} rows) ===")
for r in null_fp_paid[:30]:
    handle = r['handle']
    matches = excel_by_handle.get(handle.lower(), [])
    excel_fps = list({str(e['final_price']) for e in matches if e['final_price'] is not None})
    print(f"  id={r['id']:4d} handle={handle:<25} platform={r['platform']:<12} excel_fp={excel_fps}")

# ---- Check product_id NULL (online) ----
null_prod = [r for r in db_data if r['product_id'] is None and r['placement_type'] == 'online']
print(f"\n=== product_id NULL (online, {len(null_prod)} rows) ===")
for r in null_prod[:30]:
    handle = r['handle']
    matches = excel_by_handle.get(handle.lower(), [])
    excel_models = list({str(e['model']) for e in matches if e['model']})
    print(f"  id={r['id']:4d} handle={handle:<25} excel_model={excel_models}")

cur.close()
conn.close()
