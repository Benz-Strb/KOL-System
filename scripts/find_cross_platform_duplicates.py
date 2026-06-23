"""
Read-only audit: find KOL (person) rows that are likely the SAME real person
appearing under a DIFFERENT platform account (different `kols.id`, different
`kol_platforms` row, different platform_id) — candidates to merge into one
person now that `kol_platforms` lets one kol have several platforms.

Does NOT modify the database, does NOT auto-merge anything (per project rule:
never auto-merge ambiguous KOL records — a human decides). Exports an Excel
file for the team to fill in `same_person_or_not` + `notes`.

Only looks ACROSS different platform_id — same-platform duplicates were
already the subject of the earlier A1/A2/B/C/D audit rounds and are clear.

No reliable structured signal exists for this (contact_info is filled on only
1/739 rows, checked before writing this script) — fuzzy name matching is the
only option, same approach as the original check_kol_duplicates.py Finding B,
just applied across platforms instead of within one platform.
"""
import sys, io, re, unicodedata, difflib
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import psycopg2

DB_URL = "postgresql://postgres.hdrweioqqqpslsjizkci:Shd2025%21ffofo@aws-1-ap-northeast-1.pooler.supabase.com:5432/postgres"

conn = psycopg2.connect(DB_URL)
cur = conn.cursor()
cur.execute("""
    SELECT k.id AS kol_id, k.gen_name, kp.id AS kp_id, kp.handle, kp.platform_id,
           p.name AS platform, kp.follower_count, kp.profile_url
    FROM kols k
    JOIN kol_platforms kp ON kp.kol_id = k.id
    LEFT JOIN platforms p ON p.id = kp.platform_id
    ORDER BY k.id, kp.id
""")
rows = cur.fetchall()
cols = ['kol_id', 'gen_name', 'kp_id', 'handle', 'platform_id', 'platform', 'follower_count', 'profile_url']
accounts = [dict(zip(cols, r)) for r in rows]
print(f"Total kol_platforms rows: {len(accounts)} across {len({a['kol_id'] for a in accounts})} kols")

# ---- normalize text (fold mathematical bold/italic unicode -> plain ascii, strip punctuation/space) ----
def normalize_text(s):
    if not s:
        return ''
    s = unicodedata.normalize('NFKD', s)
    s = re.sub(r'[​‌‍﻿]', '', s)  # zero-width junk
    s = re.sub(r'[^\w]', '', s, flags=re.UNICODE)
    return s.lower()

SUFFIX_TOKENS = ['vdo', 'kit', 'photo', 'video']
def strip_format_suffix(norm):
    changed = True
    while changed:
        changed = False
        for tok in SUFFIX_TOKENS:
            if norm.endswith(tok) and len(norm) > len(tok):
                norm = norm[: -len(tok)]
                changed = True
    return norm

for a in accounts:
    a['norm_handle'] = normalize_text(a['handle'])
    a['norm_genname'] = normalize_text(a['gen_name'])
    a['norm_base'] = strip_format_suffix(a['norm_handle']) or strip_format_suffix(a['norm_genname'])

# ---- pairwise compare across DIFFERENT platform_id only ----
THRESHOLD = 0.72
n = len(accounts)
edges = []  # (i, j, ratio, exact_base)
for i in range(n):
    for j in range(i + 1, n):
        a, b = accounts[i], accounts[j]
        if a['kol_id'] == b['kol_id']:
            continue  # already the same person (shouldn't happen pre-merge, but safe)
        if a['platform_id'] == b['platform_id']:
            continue  # same-platform duplicates are a different, already-cleared audit
        exact_base = bool(a['norm_base'] and a['norm_base'] == b['norm_base'])
        texts_a = [t for t in (a['norm_handle'], a['norm_genname']) if t]
        texts_b = [t for t in (b['norm_handle'], b['norm_genname']) if t]
        best = 0.0
        for ta in texts_a:
            for tb in texts_b:
                best = max(best, difflib.SequenceMatcher(None, ta, tb).ratio())
        if exact_base or best >= THRESHOLD:
            edges.append((i, j, best, exact_base))

print(f"Candidate edges (fuzzy match across different platforms): {len(edges)}")

# ---- union-find: cluster connected accounts (A~B, B~C => one cluster) ----
parent = list(range(n))
def find(x):
    while parent[x] != x:
        parent[x] = parent[parent[x]]
        x = parent[x]
    return x
def union(x, y):
    rx, ry = find(x), find(y)
    if rx != ry:
        parent[rx] = ry

edge_by_pair = {}
for i, j, ratio, exact_base in edges:
    union(i, j)
    edge_by_pair[(i, j)] = (ratio, exact_base)

clusters = {}
for idx in range(n):
    clusters.setdefault(find(idx), []).append(idx)

# only clusters that actually span >1 distinct kol_id are real candidates
candidate_clusters = []
for members in clusters.values():
    kol_ids = {accounts[m]['kol_id'] for m in members}
    if len(kol_ids) > 1:
        max_ratio = max((edge_by_pair[(min(p), max(p))][0]
                          for p in [(a, b) for idx_a, a in enumerate(members) for b in members[idx_a + 1:]]
                          if (min(p), max(p)) in edge_by_pair), default=0.0)
        any_exact = any(edge_by_pair[(min(p), max(p))][1]
                         for p in [(a, b) for idx_a, a in enumerate(members) for b in members[idx_a + 1:]]
                         if (min(p), max(p)) in edge_by_pair)
        candidate_clusters.append((members, max_ratio, any_exact))

candidate_clusters.sort(key=lambda c: (-c[2], -c[1]))
print(f"\n=== Cross-platform duplicate candidates — {len(candidate_clusters)} clusters — NEEDS HUMAN REVIEW ===")
for members, max_ratio, any_exact in candidate_clusters:
    tag = 'EXACT-NAME-MATCH' if any_exact else f'fuzzy={max_ratio:.2f}'
    print(f"\n[{tag}]")
    for m in sorted(members, key=lambda m: accounts[m]['kol_id']):
        a = accounts[m]
        print(f"  kol_id={a['kol_id']:<5} platform={str(a['platform']):<10} handle={a['handle']!r:30} "
              f"gen_name={a['gen_name']!r:30} followers={a['follower_count']}")

# ---- Excel export ----
XLSX_PATH = 'kol_cross_platform_review.xlsx'
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Cross-platform duplicates'

HEADERS = ['group_no', 'match_confidence', 'kol_id', 'kol_platforms_id', 'platform', 'handle', 'gen_name',
           'follower_count', 'profile_url', 'same_person_or_not (กรอกเอง)', 'notes (กรอกเอง)']
ws.append(HEADERS)
header_fill = PatternFill('solid', fgColor='1F2937')
for cell in ws[1]:
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = header_fill
ws.freeze_panes = 'A2'

GROUP_FILLS = [PatternFill('solid', fgColor='F3F4F6'), PatternFill('solid', fgColor='FFFFFF')]
thin = Side(style='thin', color='D1D5DB')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

row_i = 2
for gi, (members, max_ratio, any_exact) in enumerate(candidate_clusters, start=1):
    tag = 'exact_name_match' if any_exact else f'fuzzy_{max_ratio:.2f}'
    fill = GROUP_FILLS[gi % 2]
    for m in sorted(members, key=lambda m: accounts[m]['kol_id']):
        a = accounts[m]
        ws.append([gi, tag, a['kol_id'], a['kp_id'], a['platform'], a['handle'], a['gen_name'],
                   a['follower_count'], a['profile_url'], '', ''])
        for cell in ws[row_i]:
            cell.fill = fill
            cell.border = border
        row_i += 1

dv = DataValidation(type='list', formula1='"เป็นคนเดียวกัน,เป็นคนละคน"', allow_blank=True)
ws.add_data_validation(dv)
dv.add(f'J2:J{row_i - 1}')

widths = [9, 16, 8, 16, 10, 28, 26, 14, 45, 22, 30]
for col, w in zip('ABCDEFGHIJK', widths):
    ws.column_dimensions[col].width = w

wb.save(XLSX_PATH)
print(f"\nExcel exported: {XLSX_PATH} ({len(candidate_clusters)} groups, {row_i - 2} rows)")

cur.close()
conn.close()
