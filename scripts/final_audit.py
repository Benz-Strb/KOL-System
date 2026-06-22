"""
Final targeted audit to find ALL fixable NULLs.
Uses sequential matching by (handle, platform, campaign) with fallback to (handle, platform).
Also checks payment_type mismatches (paid in DB but Free/Barter in Excel).
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
import psycopg2
import re
from decimal import Decimal, InvalidOperation
from collections import defaultdict

EXCEL_PATH = r'C:\Users\USER-SHD-046\Downloads\Copy of 2026 Dreame Plan&Summary Update.xlsx'
DB_URL = "postgresql://postgres.hdrweioqqqpslsjizkci:Shd2025%21ffofo@aws-1-ap-northeast-1.pooler.supabase.com:5432/postgres"

def normalize_campaign(val):
    if val is None: return ''
    try:
        f = float(val)
        s = f'{round(f, 2):.2f}'.rstrip('0').rstrip('.')
        return s
    except: return str(val).strip()

def parse_follower(val):
    if val is None: return None
    s = str(val).strip().upper().replace(',', '')
    if s in ('', 'NONE', '-', 'N/A', '0', '0.0'): return None
    try:
        m = re.match(r'^([\d.]+)\s*([KMB]?)$', s)
        if m:
            num = float(m.group(1))
            mult = {'K': 1000, 'M': 1_000_000, 'B': 1_000_000_000, '': 1}[m.group(2)]
            v = int(num * mult)
            return v if v > 0 else None
    except: pass
    return None

def parse_price(val):
    if val is None: return None
    s = str(val).strip().upper()
    if s in ('FREE', 'BARTER', '-', '', 'N/A', 'NONE', '0', '0.0'): return None
    try:
        d = Decimal(s.replace(',', ''))
        return d if d > 0 else None
    except InvalidOperation: return None

def infer_payment_type(val):
    if val is None: return 'paid'
    s = str(val).strip().upper()
    if s == 'FREE': return 'free'
    if s == 'BARTER': return 'barter'
    return 'paid'

# ---- Excel load ----
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws = wb['KOL Performance 2026']
merge_map = {}
for mr in ws.merged_cells.ranges:
    tl = ws.cell(mr.min_row, mr.min_col)
    for row in range(mr.min_row, mr.max_row + 1):
        for col in range(mr.min_col, mr.max_col + 1):
            merge_map[(row, col)] = tl.value
def gv(row, col):
    return merge_map.get((row, col), ws.cell(row, col).value)

excel_rows = []
for r in range(3, ws.max_row + 1):
    kol = gv(r, 6)
    platform = gv(r, 3)
    if not kol or not platform: continue
    kol_str = str(kol).strip()
    plat_str = str(platform).strip()
    if kol_str == plat_str: continue
    low = kol_str.lower()
    if low.startswith('total') or low.startswith('รวม') or low.startswith('sum'): continue
    fp_raw = gv(r, 11)
    c_norm = normalize_campaign(gv(r, 1))
    excel_rows.append({
        'excel_row': r,
        'campaign': c_norm,
        'platform': plat_str.lower(),
        'kol_handle': kol_str,
        'kol_lower': kol_str.lower(),
        'follower_int': parse_follower(gv(r, 9)),
        'final_price': parse_price(fp_raw),
        'payment_type': infer_payment_type(fp_raw),
        'pay_amount': parse_price(gv(r, 12)),
        'pay_amount_raw': gv(r, 12),
        'fp_raw': str(fp_raw or '').strip().upper(),
    })

# Build lookup maps
excel_by_3key = defaultdict(list)  # (handle, platform, campaign) -> list
excel_by_2key = defaultdict(list)  # (handle, platform) -> list
for er in excel_rows:
    k3 = (er['kol_lower'], er['platform'], er['campaign'])
    k2 = (er['kol_lower'], er['platform'])
    excel_by_3key[k3].append(er)
    excel_by_2key[k2].append(er)

# ---- DB ----
conn = psycopg2.connect(DB_URL)
cur = conn.cursor()
cur.execute("""
    SELECT p.id, k.handle, plt.name, c.code,
           p.follower_at_time, p.final_price, p.pay_amount,
           p.payment_type, p.placement_type, p.status
    FROM placements p
    JOIN kols k ON k.id = p.kol_id
    JOIN platforms plt ON plt.id = p.platform_id
    LEFT JOIN campaigns c ON c.id = p.campaign_id
    ORDER BY p.id
""")
db_data = [{'id': r[0], 'handle': r[1], 'handle_lower': r[1].lower(),
             'platform': r[2].lower(), 'campaign_code': r[3] or '',
             'follower_at_time': r[4], 'final_price': r[5], 'pay_amount': r[6],
             'payment_type': r[7], 'placement_type': r[8], 'status': r[9]}
           for r in cur.fetchall()]

# ---- Find all fixable follower NULLs ----
print("=== FOLLOWER_AT_TIME FIXES ===")
follower_fixes = []

# Track used Excel rows to avoid double-matching
used_excel_rows = set()

for db_r in db_data:
    if db_r['follower_at_time'] is not None: continue
    k3 = (db_r['handle_lower'], db_r['platform'], db_r['campaign_code'])
    k2 = (db_r['handle_lower'], db_r['platform'])

    # Try 3-key match first
    candidates = [e for e in excel_by_3key.get(k3, [])
                  if e['follower_int'] is not None and e['excel_row'] not in used_excel_rows]
    if not candidates:
        # Fall back to 2-key
        candidates = [e for e in excel_by_2key.get(k2, [])
                      if e['follower_int'] is not None and e['excel_row'] not in used_excel_rows]

    unique_vals = list({c['follower_int'] for c in candidates})
    if len(unique_vals) == 1:
        best = candidates[0]
        used_excel_rows.add(best['excel_row'])
        follower_fixes.append({
            'db_id': db_r['id'],
            'handle': db_r['handle'],
            'platform': db_r['platform'],
            'campaign': db_r['campaign_code'],
            'follower': unique_vals[0],
            'excel_row': best['excel_row'],
        })
    elif len(unique_vals) > 1:
        print(f"  AMBIGUOUS id={db_r['id']:4d} {db_r['handle']:<30} vals={unique_vals}")

print(f"\nTotal follower fixes: {len(follower_fixes)}")
for f in follower_fixes:
    print(f"  UPDATE id={f['db_id']:4d}: follower_at_time={f['follower']}  "
          f"-- {f['handle']}/{f['platform']}/campaign={f['campaign']} (excel row {f['excel_row']})")

# ---- Find all fixable final_price NULLs (paid) ----
print("\n=== FINAL_PRICE FIXES (paid, NULL) ===")
price_fixes = []
price_type_fixes = []  # payment_type mismatch: DB=paid but Excel=free/barter

used_excel_fp = set()

for db_r in db_data:
    if db_r['payment_type'] != 'paid': continue
    k3 = (db_r['handle_lower'], db_r['platform'], db_r['campaign_code'])
    k2 = (db_r['handle_lower'], db_r['platform'])

    # Check payment_type mismatch
    all_matches = (excel_by_3key.get(k3, []) or excel_by_2key.get(k2, []))
    if all_matches:
        excel_ptypes = list({e['payment_type'] for e in all_matches})
        excel_fps = list({e['fp_raw'] for e in all_matches})

    if db_r['final_price'] is not None: continue

    candidates = [e for e in excel_by_3key.get(k3, [])
                  if e['final_price'] is not None and e['excel_row'] not in used_excel_fp]
    if not candidates:
        candidates = [e for e in excel_by_2key.get(k2, [])
                      if e['final_price'] is not None and e['excel_row'] not in used_excel_fp]

    unique_vals = list({str(c['final_price']) for c in candidates})
    if len(unique_vals) == 1:
        best = candidates[0]
        used_excel_fp.add(best['excel_row'])
        price_fixes.append({
            'db_id': db_r['id'],
            'handle': db_r['handle'],
            'platform': db_r['platform'],
            'campaign': db_r['campaign_code'],
            'price': best['final_price'],
            'excel_row': best['excel_row'],
        })
    elif len(unique_vals) > 1:
        print(f"  AMBIGUOUS id={db_r['id']:4d} {db_r['handle']:<30} vals={unique_vals}")

print(f"\nTotal final_price fixes: {len(price_fixes)}")
for f in price_fixes:
    print(f"  UPDATE id={f['db_id']:4d}: final_price={f['price']}  "
          f"-- {f['handle']}/{f['platform']}/campaign={f['campaign']} (excel row {f['excel_row']})")

# ---- Check payment_type mismatches ----
print("\n=== PAYMENT_TYPE MISMATCHES (DB=paid but Excel=barter/free) ===")
pt_mismatches = []
for db_r in db_data:
    if db_r['payment_type'] != 'paid': continue
    k3 = (db_r['handle_lower'], db_r['platform'], db_r['campaign_code'])
    k2 = (db_r['handle_lower'], db_r['platform'])
    candidates = excel_by_3key.get(k3) or excel_by_2key.get(k2) or []
    if candidates:
        excel_ptypes = {c['payment_type'] for c in candidates}
        if excel_ptypes == {'barter'}:
            pt_mismatches.append((db_r, 'barter', [c['fp_raw'] for c in candidates[:2]]))
        elif excel_ptypes == {'free'}:
            pt_mismatches.append((db_r, 'free', [c['fp_raw'] for c in candidates[:2]]))

print(f"Total payment_type mismatches: {len(pt_mismatches)}")
for db_r, correct_type, raw_vals in pt_mismatches:
    print(f"  id={db_r['id']:4d} {db_r['handle']:<30} {db_r['platform']:<12} "
          f"DB=paid -> should be {correct_type} (excel: {raw_vals})")

# ---- pay_amount fixes ----
print("\n=== PAY_AMOUNT FIXES (paid, NULL pay_amount but Excel has value) ===")
pa_fixes = []
used_excel_pa = set()
for db_r in db_data:
    if db_r['pay_amount'] is not None: continue
    if db_r['payment_type'] != 'paid': continue
    k3 = (db_r['handle_lower'], db_r['platform'], db_r['campaign_code'])
    k2 = (db_r['handle_lower'], db_r['platform'])
    candidates = [e for e in (excel_by_3key.get(k3) or excel_by_2key.get(k2) or [])
                  if e['pay_amount'] is not None and e['excel_row'] not in used_excel_pa]
    unique_vals = list({str(c['pay_amount']) for c in candidates})
    if len(unique_vals) == 1:
        best = candidates[0]
        used_excel_pa.add(best['excel_row'])
        pa_fixes.append({
            'db_id': db_r['id'],
            'handle': db_r['handle'],
            'platform': db_r['platform'],
            'campaign': db_r['campaign_code'],
            'pay_amount': best['pay_amount'],
            'excel_row': best['excel_row'],
        })

print(f"Total pay_amount fixes: {len(pa_fixes)}")
for f in pa_fixes[:20]:
    print(f"  UPDATE id={f['db_id']:4d}: pay_amount={f['pay_amount']}  "
          f"-- {f['handle']}/{f['platform']}")

cur.close()
conn.close()
