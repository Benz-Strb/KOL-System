"""
Read-only export: the 15 A1 groups that did NOT fit the "1 base + suffix
variant(s), same platform" pattern (already merged separately, see
run_a1_merge.py + kol_merge_backup_a1_*.json). These need a human decision —
no clear base row, multiple base-looking rows, or a cross-platform url
collision. Exported for the team to fill in.

Does NOT modify the database.
"""
import sys, io, re, unicodedata
from urllib.parse import urlparse, parse_qs
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import psycopg2

DB_URL = "postgresql://postgres.hdrweioqqqpslsjizkci:Shd2025%21ffofo@aws-1-ap-northeast-1.pooler.supabase.com:5432/postgres"

conn = psycopg2.connect(DB_URL)
cur = conn.cursor()
cur.execute("""
    SELECT k.id, k.handle, k.gen_name, k.profile_url, k.follower_count, p.name AS platform,
           (SELECT COUNT(*) FROM placements pl WHERE pl.kol_id = k.id) AS placement_count,
           (SELECT COUNT(*) FROM kol_commercial_terms t WHERE t.kol_id = k.id) AS terms_count,
           (SELECT COUNT(*) FROM kol_samples s WHERE s.kol_id = k.id) AS samples_count
    FROM kols k
    LEFT JOIN platforms p ON p.id = k.platform_id
    ORDER BY k.id
""")
cols = ['id', 'handle', 'gen_name', 'profile_url', 'follower_count', 'platform',
        'placement_count', 'terms_count', 'samples_count']
kols = [dict(zip(cols, r)) for r in cur.fetchall()]
print(f"Total KOL rows: {len(kols)}")

def normalize_text(s):
    if not s: return ''
    s = unicodedata.normalize('NFKD', s)
    s = re.sub(r'[​‌‍﻿]', '', s)
    s = re.sub(r'[^\w]', '', s, flags=re.UNICODE)
    return s.lower()

def normalize_url(u):
    if not u: return None
    u = u.strip()
    p = urlparse(u if '://' in u else 'https://' + u)
    netloc = p.netloc.lower().replace('www.', '').replace('m.facebook.com', 'facebook.com')
    path = p.path.rstrip('/')
    qs = parse_qs(p.query)
    keep_q = ''
    if 'id' in qs and 'profile.php' in path:
        keep_q = f"?id={qs['id'][0]}"
    return f"{netloc}{path}{keep_q}"

POST_PATTERNS = {
    'facebook':  [r'/posts/', r'/videos/', r'/photos/', r'/watch/?\?', r'/reel/', r'permalink\.php', r'story_fbid', r'/video\.php'],
    'instagram': [r'/p/', r'/reel/', r'/tv/'],
    'tiktok':    [r'/video/\d+'],
    'youtube':   [r'/watch\?v=', r'/shorts/'],
    'lemon8':    [r'/p/'], 'lamon8': [r'/p/'],
}
def classify_url(platform, url):
    if not url: return 'none'
    plat = (platform or '').lower()
    for pat in POST_PATTERNS.get(plat, []):
        if re.search(pat, url, re.IGNORECASE): return 'post'
    return 'profile'

for k in kols:
    k['norm_url'] = normalize_url(k['profile_url'])
    k['url_kind'] = classify_url(k['platform'], k['profile_url'])
    k['norm_handle'] = normalize_text(k['handle'])
    k['norm_genname'] = normalize_text(k['gen_name'])

SUFFIX_TOKENS = ['vdo', 'kit', 'photo', 'video']
def strip_suffix(norm):
    changed = True
    while changed:
        changed = False
        for tok in SUFFIX_TOKENS:
            if norm.endswith(tok) and len(norm) > len(tok):
                norm = norm[:-len(tok)]; changed = True
    return norm
def has_suffix(norm):
    return any(norm.endswith(tok) and len(norm) > len(tok) for tok in SUFFIX_TOKENS)

def group_kind(group):
    bases = set()
    for k in group:
        base = strip_suffix(k['norm_handle']) or strip_suffix(k['norm_genname'])
        bases.add(base)
    return 'variant_pattern' if len(bases) <= 1 and all(bases) else 'needs_review'

by_url = {}
for k in kols:
    if k['url_kind'] != 'profile' or not k['norm_url']: continue
    by_url.setdefault(k['norm_url'], []).append(k)

groups_all = [(u, g) for u, g in by_url.items() if len(g) > 1]
groups_variant = [(u, g) for u, g in groups_all if group_kind(g) == 'variant_pattern']

ambiguous = []
for url, group in groups_variant:
    no_suffix = [k for k in group if not (has_suffix(k['norm_handle']) or has_suffix(k['norm_genname']))]
    suffix = [k for k in group if k not in no_suffix]
    platforms = set(k['platform'] for k in group)
    if len(no_suffix) == 1 and len(suffix) >= 1 and len(platforms) == 1:
        continue  # already merged by run_a1_merge.py
    ambiguous.append((url, group))

print(f"Remaining ambiguous groups: {len(ambiguous)}")

XLSX_PATH = 'kol_duplicates_A1_remaining_review.xlsx'
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'รอคุย'
HEADERS = ['group_no', 'shared_profile_url', 'kol_id', 'platform', 'handle', 'gen_name',
           'follower_count', 'placements', 'commercial_terms', 'samples',
           'keep_separate_or_merge (กรอกเอง)', 'merge_into_id (กรอกเอง ถ้าเลือกรวม)', 'notes (กรอกเอง)']
ws.append(HEADERS)
header_fill = PatternFill('solid', fgColor='1F2937')
for cell in ws[1]:
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = header_fill
ws.freeze_panes = 'A2'

GROUP_FILLS = [PatternFill('solid', fgColor='F3F4F6'), PatternFill('solid', fgColor='FFFFFF')]
thin = Side(style='thin', color='D1D5DB')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

row_i = 2
for gi, (url, group) in enumerate(ambiguous, start=1):
    group_sorted = sorted(group, key=lambda k: (k['follower_count'] is None, -(k['follower_count'] or 0)))
    fill = GROUP_FILLS[gi % 2]
    for k in group_sorted:
        ws.append([gi, url, k['id'], k['platform'], k['handle'], k['gen_name'],
                   k['follower_count'], k['placement_count'], k['terms_count'], k['samples_count'],
                   '', '', ''])
        for cell in ws[row_i]:
            cell.fill = fill
            cell.border = border
        row_i += 1

dv = DataValidation(type='list', formula1='"เก็บแยกไว้ตามเดิม,รวมเข้า id อื่น"', allow_blank=True)
ws.add_data_validation(dv)
dv.add(f'K2:K{row_i - 1}')

widths = [9, 35, 8, 10, 30, 26, 14, 11, 16, 9, 28, 18, 30]
for col, w in zip('ABCDEFGHIJKLM', widths):
    ws.column_dimensions[col].width = w

wb.save(XLSX_PATH)
print(f"Excel exported: {XLSX_PATH} ({len(ambiguous)} groups, {row_i - 2} rows)")

cur.close()
conn.close()
