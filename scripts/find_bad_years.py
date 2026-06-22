"""
ค้นหาทุก cell ใน col O ที่ปีผิด pattern '0206', '0225' ฯลฯ (02XX → 20XX)
และ fix publication_date ใน DB ที่ยัง NULL
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
import psycopg2
import re
from datetime import datetime, date
from collections import defaultdict

EXCEL_PATH = r'C:\Users\USER-SHD-046\Downloads\Copy of 2026 Dreame Plan&Summary Update.xlsx'
DB_URL = "postgresql://postgres.hdrweioqqqpslsjizkci:Shd2025%21ffofo@aws-1-ap-northeast-1.pooler.supabase.com:5432/postgres"

def normalize_campaign(val):
    if val is None: return ''
    try:
        f = float(val)
        return f'{round(f,2):.2f}'.rstrip('0').rstrip('.')
    except: return str(val).strip()

def is_url(val):
    return bool(val) and str(val).strip().startswith('http')

SYSTEM_YEAR = 2026  # ข้อมูลทั้งหมดในระบบนี้เป็นปี 2026

def fix_bad_year(raw):
    """
    ถ้า year ที่ parse ได้ไม่อยู่ใน 2020-2030 → fix เป็น SYSTEM_YEAR (2026)
    ตีความ text เป็น DD/MM (Thai format)
    คืน (date, note) หรือ (None, None) ถ้าปีปกติหรือ parse ไม่ได้
    """
    if raw is None or is_url(raw): return None, None

    if isinstance(raw, (datetime, date)):
        d = raw if isinstance(raw, date) else raw.date()
        if 2020 <= d.year <= 2030: return None, None  # ปกติ ไม่ต้องแก้
        try:
            return date(SYSTEM_YEAR, d.month, d.day), f"datetime year {d.year}→{SYSTEM_YEAR}"
        except: return None, None

    s = str(raw).strip()
    if not s: return None, None

    # Match DD/MM/YYYY (Thai format)
    m = re.match(r'^(\d{1,2})/(\d{1,2})/(\d+)$', s)
    if m:
        d_str, mo_str, y_str = m.groups()
        y = int(y_str)
        if 2020 <= y <= 2030: return None, None  # ปีถูกแล้ว
        try:
            return date(SYSTEM_YEAR, int(mo_str), int(d_str)), f"text year '{y_str}'→{SYSTEM_YEAR}"
        except:
            try:  # ถ้า mo/day สลับกัน
                return date(SYSTEM_YEAR, int(d_str), int(mo_str)), f"text year '{y_str}'→{SYSTEM_YEAR} (swapped)"
            except: pass

    return None, None

# Load Excel
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws = wb['KOL Performance 2026']
merge_map = {}
merge_top_left = {}
for mr in ws.merged_cells.ranges:
    tl_row, tl_col = mr.min_row, mr.min_col
    for row in range(mr.min_row, mr.max_row+1):
        for col in range(mr.min_col, mr.max_col+1):
            merge_map[(row,col)] = ws.cell(tl_row,tl_col).value
            merge_top_left[(row,col)] = (tl_row, tl_col)

def gv(r,c): return merge_map.get((r,c), ws.cell(r,c).value)

def get_col_o_raw(row):
    ar, _ = merge_top_left.get((row,15),(row,15))
    return ws.cell(ar,15).value

# สแกนทุก row
bad_rows = []
excel_rows = []

for r in range(3, ws.max_row+1):
    kol = gv(r,6); plat = gv(r,3)
    if not kol or not plat: continue
    ks, ps = str(kol).strip(), str(plat).strip()
    if ks==ps: continue
    if ks.lower().startswith(('total','รวม','sum')): continue
    camp = normalize_campaign(gv(r,1))
    raw = get_col_o_raw(r)

    fixed, note = fix_bad_year(raw)
    if fixed:
        bad_rows.append({'excel_row':r,'handle':ks,'platform':ps,
                         'campaign':camp,'raw':raw,'fixed':fixed,'note':note})

    excel_rows.append({'excel_row':r,'handle':ks,'platform':ps,
                       'campaign':camp,'pub_date':fixed})

print(f"=== Excel col O ที่ปีผิด (02XX→20XX): {len(bad_rows)} rows ===\n")
# Group เพื่อแสดงกระชับ
seen_raw = {}
for b in bad_rows:
    key = str(b['raw'])
    seen_raw.setdefault(key, []).append(b['excel_row'])

for raw_val, rows in seen_raw.items():
    # หา fixed date
    fd = next(b['fixed'] for b in bad_rows if str(b['raw'])==raw_val)
    print(f"  raw='{raw_val}'  →  {fd}  (Excel rows: {rows})")

# DB
conn = psycopg2.connect(DB_URL)
cur = conn.cursor()
cur.execute("""
    SELECT p.id, k.handle, plt.name, c.code, p.publication_date
    FROM placements p JOIN kols k ON k.id=p.kol_id
    LEFT JOIN platforms plt ON plt.id=p.platform_id
    LEFT JOIN campaigns c ON c.id=p.campaign_id
    ORDER BY p.id
""")
db_by_3key = defaultdict(list)
for r in cur.fetchall():
    key = (r[1].lower(),(r[2] or '').lower(),r[3] or '')
    db_by_3key[key].append({'id':r[0],'handle':r[1],'platform':r[2] or '',
                             'campaign':r[3] or '','publication_date':r[4]})

excel_used = set()
fixes = []
for ep in excel_rows:
    if ep['pub_date'] is None: continue
    k3 = (ep['handle'].lower(), ep['platform'].lower(), ep['campaign'])
    candidates = [d for d in db_by_3key.get(k3,[]) if d['id'] not in excel_used]
    if not candidates: continue
    dr = candidates[0]
    excel_used.add(dr['id'])
    if dr['publication_date'] is None:
        fixes.append({**dr, 'new_date': ep['pub_date']})

# กรองเฉพาะที่มาจาก bad_rows
bad_excel_set = {b['excel_row'] for b in bad_rows}
# re-match เฉพาะ bad rows
excel_used2 = set()
bad_fixes = []
for ep in excel_rows:
    if ep['excel_row'] not in bad_excel_set or ep['pub_date'] is None: continue
    k3 = (ep['handle'].lower(), ep['platform'].lower(), ep['campaign'])
    candidates = [d for d in db_by_3key.get(k3,[]) if d['id'] not in excel_used2]
    if not candidates: continue
    dr = candidates[0]
    excel_used2.add(dr['id'])
    if dr['publication_date'] is None:
        bad_fixes.append({**dr, 'new_date': ep['pub_date']})

print(f"\n=== DB publication_date NULL เพราะ bad-year ใน Excel: {len(bad_fixes)} rows ===")
for f in bad_fixes:
    print(f"  id={f['id']:4d}  [{f['handle'][:25]}/{f['platform']}/camp={f['campaign']}]  → {f['new_date']}")

if bad_fixes:
    sql_path = r'D:\internship\KOL_management\kol-system\scripts\fix_bad_year_pubdates.sql'
    with open(sql_path, 'w', encoding='utf-8') as fout:
        fout.write("-- publication_date fixes: bad year 02XX→20XX in Excel col O\n\n")
        for f in bad_fixes:
            fout.write(f"UPDATE placements SET publication_date = '{f['new_date']}' WHERE id = {f['id']};\n")
    print(f"\nSQL ({len(bad_fixes)} statements) → {sql_path}")
else:
    print("ไม่มีแถวเพิ่มเติมที่ต้องแก้")

cur.close()
conn.close()
