import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl
import psycopg2
from datetime import datetime, date

EXCEL_PATH = r'C:\Users\USER-SHD-046\Downloads\Copy of 2026 Dreame Plan&Summary Update.xlsx'
DB_URL = "postgresql://postgres.hdrweioqqqpslsjizkci:Shd2025%21ffofo@aws-1-ap-northeast-1.pooler.supabase.com:5432/postgres"

wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws = wb['KOL Performance 2026']

merge_top_left = {}
for mr in ws.merged_cells.ranges:
    tl_row, tl_col = mr.min_row, mr.min_col
    for row in range(mr.min_row, mr.max_row+1):
        for col in range(mr.min_col, mr.max_col+1):
            merge_top_left[(row, col)] = (tl_row, tl_col)

# ดู raw cell value ของ col O rows 718-722
print("=== col O raw (merge-aware) rows 718-722 ===")
for r in range(718, 723):
    ar, _ = merge_top_left.get((r, 15), (r, 15))
    raw = ws.cell(ar, 15).value
    print(f"  row={r}  top_left_row={ar}  raw='{raw}'  type={type(raw).__name__}")

# DB: publication_date ของ id 694-696
conn = psycopg2.connect(DB_URL)
cur = conn.cursor()
cur.execute("""
    SELECT p.id, k.handle, p.publication_date, p.post_url
    FROM placements p JOIN kols k ON k.id = p.kol_id
    WHERE p.id IN (694, 695, 696)
    ORDER BY p.id
""")
print("\n=== DB id 694-696 ===")
for r in cur.fetchall():
    print(f"  id={r[0]}  handle={r[1]}  publication_date={r[2]}  post_url={r[3]}")

cur.close()
conn.close()
