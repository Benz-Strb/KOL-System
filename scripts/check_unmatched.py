"""
ตรวจสอบ 4 unmatched rows ใหม่ และดูค่า Excel ของ id=837 และ id=461
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
import psycopg2
import re
from decimal import Decimal, InvalidOperation
from collections import defaultdict

EXCEL_PATH = r'C:\Users\USER-SHD-046\Downloads\Copy of 2026 Dreame Plan&Summary Update.xlsx'
DB_URL = "postgresql://postgres.hdrweioqqqpslsjizkci:Shd2025%21ffofo@aws-1-ap-northeast-1.pooler.supabase.com:5432/postgres"

def normalize_campaign(val):
    if val is None: return ''
    try:
        f = float(val)
        s = f'{round(f, 2):.2f}'.rstrip('0').rstrip('.')
        return s
    except:
        return str(val).strip()

def parse_price(val):
    if val is None: return None
    s = str(val).strip().upper()
    if s in ('FREE', 'BARTER', '-', '', 'N/A', 'NONE', '0', '0.0'): return None
    try:
        d = Decimal(str(s).replace(',', ''))
        return d if d > 0 else None
    except InvalidOperation:
        return None

wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws = wb['KOL Performance 2026']

merge_map = {}
for mr in ws.merged_cells.ranges:
    tl = ws.cell(mr.min_row, mr.min_col)
    for row in range(mr.min_row, mr.max_row + 1):
        for col in range(mr.min_col, mr.max_col + 1):
            merge_map[(row, col)] = tl.value

def gv(row, col):
    return merge_map.get((row, col), ws.cell(row, col).value)

# ---- ค้นหา 4 handles ใน Excel ----
targets = [
    'ป๊าโทนี่ที่เป็นพ่อแมว',
    'ไปนู่นไปนี่ฉันจะตีก้นเธอ',
    'แม่น้องมาเหนือ&มานนท์',
]
print("=== ค้นหา handles ที่ไม่ match ใน Excel ===")
for r in range(3, ws.max_row + 1):
    kol = gv(r, 6)
    if kol is None: continue
    kol_str = str(kol).strip()
    for t in targets:
        if t.lower() in kol_str.lower() or kol_str.lower() in t.lower():
            camp = normalize_campaign(gv(r, 1))
            plat = gv(r, 3)
            pic  = gv(r, 2)
            fp   = gv(r, 11)
            print(f"  row={r:4d}  kol='{kol_str}'  platform={plat}  campaign={camp}  PIC={pic}  final_price={fp}")

# ค้น substring แบบกว้างขึ้น
print("\n=== ค้น 'โทนี่' / 'ไปนู่น' / 'มาเหนือ' ใน Excel (substring) ===")
for r in range(3, ws.max_row + 1):
    kol = gv(r, 6)
    if kol is None: continue
    kol_str = str(kol).strip()
    for sub in ['โทนี่', 'ไปนู่น', 'มาเหนือ']:
        if sub in kol_str:
            camp = normalize_campaign(gv(r, 1))
            plat = gv(r, 3)
            print(f"  row={r:4d}  kol='{kol_str}'  platform={plat}  campaign={camp}")

# ---- ดูค่า Excel ของ id=837 (supanutmakpramool, TikTok, camp=5.5) ----
print("\n=== Excel rows สำหรับ 'supanutmakpramool' (ตรวจ id=837) ===")
for r in range(3, ws.max_row + 1):
    kol = gv(r, 6)
    if kol is None: continue
    kol_str = str(kol).strip()
    if 'supanut' in kol_str.lower():
        camp = normalize_campaign(gv(r, 1))
        plat = gv(r, 3)
        pic  = gv(r, 2)
        fp   = gv(r, 11)
        pa   = gv(r, 12)
        fol  = gv(r, 9)
        post = ws.cell(r, 15).value
        print(f"  row={r:4d}  kol='{kol_str}'  platform={plat}  campaign={camp}")
        print(f"    PIC={pic}  follower={fol}  final_price={fp}  pay_amount={pa}  post_url={post}")

# ---- ดูค่า Excel ของ id=461 (โปรป้ายเหลือง, Facebook, camp=3.3) ----
print("\n=== Excel rows สำหรับ 'โปรป้ายเหลือง' (ตรวจ id=461) ===")
for r in range(3, ws.max_row + 1):
    kol = gv(r, 6)
    if kol is None: continue
    kol_str = str(kol).strip()
    if 'ป้ายเหลือง' in kol_str:
        camp = normalize_campaign(gv(r, 1))
        plat = gv(r, 3)
        pic  = gv(r, 2)
        fp   = gv(r, 11)
        pa   = gv(r, 12)
        fol  = gv(r, 9)
        print(f"  row={r:4d}  kol='{kol_str}'  platform={plat}  campaign={camp}")
        print(f"    PIC={pic}  follower={fol}  final_price={fp}  pay_amount={pa}")

# ---- ดูค่า DB ของ id=461, 837 ----
print("\n=== DB values สำหรับ id=461 และ 837 ===")
conn = psycopg2.connect(DB_URL)
cur = conn.cursor()
cur.execute("""
    SELECT p.id, k.handle, plt.name, c.code,
           p.final_price, p.pay_amount, p.payment_type, p.follower_at_time, p.status
    FROM placements p
    JOIN kols k ON k.id = p.kol_id
    LEFT JOIN platforms plt ON plt.id = p.platform_id
    LEFT JOIN campaigns c ON c.id = p.campaign_id
    WHERE p.id IN (461, 837)
""")
for r in cur.fetchall():
    print(f"  id={r[0]}  handle={r[1]}  platform={r[2]}  campaign={r[3]}")
    print(f"    final_price={r[4]}  pay_amount={r[5]}  payment_type={r[6]}  follower={r[7]}  status={r[8]}")

# ---- ดู DB rows ของ 4 handles ใหม่ ----
print("\n=== DB rows ของ 4 handles ที่ไม่ match ===")
cur.execute("""
    SELECT p.id, k.handle, plt.name, c.code, p.final_price, p.pay_amount,
           p.payment_type, p.follower_at_time, p.status, p.created_at
    FROM placements p
    JOIN kols k ON k.id = p.kol_id
    LEFT JOIN platforms plt ON plt.id = p.platform_id
    LEFT JOIN campaigns c ON c.id = p.campaign_id
    WHERE k.handle IN ('ป๊าโทนี่ที่เป็นพ่อแมว', 'ไปนู่นไปนี่ฉันจะตีก้นเธอ', 'แม่น้องมาเหนือ&มานนท์')
    ORDER BY p.id
""")
for r in cur.fetchall():
    print(f"  id={r[0]:4d}  handle={r[1]}  platform={r[2]}  campaign={r[3]}")
    print(f"    final_price={r[4]}  pay_amount={r[5]}  payment_type={r[6]}  follower={r[7]}  status={r[8]}  created={r[9]}")

cur.close()
conn.close()
