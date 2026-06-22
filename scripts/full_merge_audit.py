"""
Comprehensive merge-cell audit for ALL columns in placements.
Compares Excel (merge-cell aware) vs DB and generates SQL to fix discrepancies.

Special cases:
- Rows with NULL platform_id: matched by (handle + campaign) since platform unknown
- post_url: only checked for status='posted' rows (planned = not yet posted = NULL is normal)
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
import psycopg2
import re
from decimal import Decimal, InvalidOperation
from collections import defaultdict

EXCEL_PATH = r'C:\Users\USER-SHD-046\Downloads\Copy of 2026 Dreame Plan&Summary Update.xlsx'
DB_URL = "postgresql://postgres.hdrweioqqqpslsjizkci:Shd2025%21ffofo@aws-1-ap-northeast-1.pooler.supabase.com:5432/postgres"

# ---- helpers ----
def normalize_campaign(val):
    if val is None: return ''
    try:
        f = float(val)
        s = f'{round(f, 2):.2f}'.rstrip('0').rstrip('.')
        return s
    except:
        return str(val).strip()

def parse_follower(val):
    if val is None: return None
    s = str(val).strip().upper().replace(',', '')
    if s in ('', 'NONE', '-', 'N/A', '0', '0.0'): return None
    try:
        m = re.match(r'^([\d.]+)\s*([KMB]?)$', s)
        if m:
            num = float(m.group(1))
            mult = {'K': 1000, 'M': 1_000_000, 'B': 1_000_000_000, '': 1}[m.group(2)]
            v = int(num * mult)
            return v if v > 0 else None
    except:
        pass
    return None

def parse_price(val):
    if val is None: return None
    s = str(val).strip().upper()
    if s in ('FREE', 'BARTER', '-', '', 'N/A', 'NONE', '0', '0.0'): return None
    try:
        d = Decimal(s.replace(',', ''))
        return d if d > 0 else None
    except InvalidOperation:
        return None

def infer_payment_type(val):
    if val is None: return 'paid'
    s = str(val).strip().upper()
    if s == 'FREE': return 'free'
    if s == 'BARTER': return 'barter'
    return 'paid'

def is_url(val):
    if val is None: return False
    s = str(val).strip()
    return s.startswith('http://') or s.startswith('https://')

# ---- Load Excel (merge-cell aware for cols A-F) ----
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws = wb['KOL Performance 2026']

merge_map = {}
for mr in ws.merged_cells.ranges:
    tl = ws.cell(mr.min_row, mr.min_col)
    for row in range(mr.min_row, mr.max_row + 1):
        for col in range(mr.min_col, mr.max_col + 1):
            merge_map[(row, col)] = tl.value

def gv(row, col):
    return merge_map.get((row, col), ws.cell(row, col).value)

excel_rows = []
for r in range(3, ws.max_row + 1):
    kol = gv(r, 6)
    platform = gv(r, 3)
    if not kol or not platform:
        continue
    kol_str = str(kol).strip()
    plat_str = str(platform).strip()
    if kol_str == plat_str:
        continue
    low = kol_str.lower()
    if low.startswith('total') or low.startswith('รวม') or low.startswith('sum'):
        continue

    fp_raw = gv(r, 11)
    post_raw = ws.cell(r, 15).value  # col O — raw (not merged), URL or date
    post_url = str(post_raw).strip() if is_url(post_raw) else None

    excel_rows.append({
        'excel_row':    r,
        'campaign':     normalize_campaign(gv(r, 1)),
        'pic':          str(gv(r, 2) or '').strip(),
        'platform':     plat_str,
        'model_code':   str(gv(r, 5) or '').strip(),
        'kol_handle':   kol_str,
        'follower_int': parse_follower(gv(r, 9)),
        'fp_raw':       str(fp_raw or '').strip().upper(),
        'final_price':  parse_price(fp_raw),
        'payment_type': infer_payment_type(fp_raw),
        'pay_amount':   parse_price(gv(r, 12)),
        'post_url':     post_url,
        'post_raw':     str(post_raw or '').strip(),  # for debugging
    })

print(f"Excel data rows: {len(excel_rows)}")

# ---- Connect DB ----
conn = psycopg2.connect(DB_URL)
cur = conn.cursor()

# Lookup tables
cur.execute("SELECT id, name FROM platforms")
plat_rows = cur.fetchall()
plat_name_to_id = {r[1].lower(): r[0] for r in plat_rows}
plat_id_to_name = {r[0]: r[1] for r in plat_rows}

cur.execute("SELECT id, full_name FROM users")
user_rows = cur.fetchall()
user_name_to_id = {r[1].lower(): r[0] for r in user_rows}
print("Users in DB:", [r[1] for r in user_rows])

cur.execute("SELECT id, model_code FROM products")
prod_rows = cur.fetchall()
prod_code_to_id = {r[1].lower(): r[0] for r in prod_rows}

# Main placements query
cur.execute("""
    SELECT p.id, k.handle, plt.name, c.code,
           p.kol_id, p.platform_id, p.campaign_id, p.product_id,
           p.person_in_charge_id, p.person_in_charge,
           p.follower_at_time, p.final_price, p.pay_amount, p.payment_type,
           p.post_url, p.placement_type, p.status
    FROM placements p
    JOIN kols k ON k.id = p.kol_id
    LEFT JOIN platforms plt ON plt.id = p.platform_id
    LEFT JOIN campaigns c ON c.id = p.campaign_id
    ORDER BY p.id
""")
db_rows = []
for r in cur.fetchall():
    db_rows.append({
        'id': r[0], 'handle': r[1], 'platform': r[2] or '', 'campaign_code': r[3] or '',
        'kol_id': r[4], 'platform_id': r[5], 'campaign_id': r[6], 'product_id': r[7],
        'person_in_charge_id': r[8], 'person_in_charge': r[9],
        'follower_at_time': r[10], 'final_price': r[11], 'pay_amount': r[12],
        'payment_type': r[13], 'post_url': r[14],
        'placement_type': r[15], 'status': r[16],
    })
print(f"DB placements: {len(db_rows)}\n")

# ---- NULL summary ----
print("=== NULL SUMMARY IN DB ===")
print(f"  platform_id NULL:          {sum(1 for r in db_rows if r['platform_id'] is None)}")
print(f"  person_in_charge_id NULL:  {sum(1 for r in db_rows if r['person_in_charge_id'] is None)}")
print(f"  product_id NULL (online):  {sum(1 for r in db_rows if r['product_id'] is None and r['placement_type'] == 'online')}")
print(f"  post_url NULL (posted):    {sum(1 for r in db_rows if r['post_url'] is None and r['status'] == 'posted')}")
print(f"  follower_at_time NULL:     {sum(1 for r in db_rows if r['follower_at_time'] is None)}")
print(f"  final_price NULL (paid):   {sum(1 for r in db_rows if r['final_price'] is None and r['payment_type'] == 'paid')}")
print(f"  pay_amount NULL (paid):    {sum(1 for r in db_rows if r['pay_amount'] is None and r['payment_type'] == 'paid')}")
print()

# ---- Build Excel lookup maps ----
excel_by_3key = defaultdict(list)  # (handle_lower, platform_lower, campaign) -> [rows]
excel_by_2key = defaultdict(list)  # (handle_lower, campaign) -> [rows]  for NULL-platform rows
for er in excel_rows:
    k3 = (er['kol_handle'].lower(), er['platform'].lower(), er['campaign'])
    k2 = (er['kol_handle'].lower(), er['campaign'])
    excel_by_3key[k3].append(er)
    excel_by_2key[k2].append(er)

excel_used = set()
fixes = {}
warnings = []

for db_r in db_rows:
    db_id = db_r['id']

    # Choose matching strategy
    if db_r['platform_id'] is None:
        # platform unknown — match by (handle + campaign) only
        k2 = (db_r['handle'].lower(), db_r['campaign_code'])
        candidates = [e for e in excel_by_2key.get(k2, []) if e['excel_row'] not in excel_used]
        match_type = '2key(no-platform)'
    else:
        k3 = (db_r['handle'].lower(), db_r['platform'].lower(), db_r['campaign_code'])
        candidates = [e for e in excel_by_3key.get(k3, []) if e['excel_row'] not in excel_used]
        match_type = '3key'

    if not candidates:
        if db_r['platform_id'] is None:
            warnings.append(f"id={db_id} [{db_r['handle']}/NULL-platform/camp={db_r['campaign_code']}]: Excel row ไม่พบ")
        continue

    ex = candidates[0]
    excel_used.add(ex['excel_row'])
    row_fixes = {}

    # 1. platform_id
    if db_r['platform_id'] is None and ex['platform']:
        pid = plat_name_to_id.get(ex['platform'].lower())
        if pid:
            row_fixes['platform_id'] = pid
        else:
            warnings.append(f"id={db_id}: platform '{ex['platform']}' ไม่พบใน platforms table")

    # 2. person_in_charge_id + person_in_charge text
    if db_r['person_in_charge_id'] is None and ex['pic']:
        uid = user_name_to_id.get(ex['pic'].lower())
        if uid:
            row_fixes['person_in_charge_id'] = uid
            row_fixes['person_in_charge'] = ex['pic']
        else:
            warnings.append(f"id={db_id}: PIC '{ex['pic']}' ไม่พบใน users table")

    # 3. product_id (online only, match by model_code)
    if db_r['product_id'] is None and ex['model_code'] and db_r['placement_type'] == 'online':
        pid = prod_code_to_id.get(ex['model_code'].lower())
        if pid:
            row_fixes['product_id'] = pid
        else:
            warnings.append(f"id={db_id}: model_code '{ex['model_code']}' ไม่พบใน products")

    # 4. post_url — only fix posted rows (planned = ยังไม่โพสต์)
    if db_r['post_url'] is None and db_r['status'] == 'posted' and ex['post_url']:
        row_fixes['post_url'] = ex['post_url']

    # 5. follower_at_time
    if db_r['follower_at_time'] is None and ex['follower_int'] is not None:
        row_fixes['follower_at_time'] = ex['follower_int']

    # 6. final_price (paid only)
    if db_r['final_price'] is None and db_r['payment_type'] == 'paid' and ex['final_price'] is not None:
        row_fixes['final_price'] = ex['final_price']

    # 7. payment_type mismatch (DB=paid, Excel=free/barter)
    if db_r['payment_type'] == 'paid' and ex['payment_type'] in ('free', 'barter'):
        row_fixes['payment_type'] = ex['payment_type']

    # 8. pay_amount (paid only)
    if db_r['pay_amount'] is None and db_r['payment_type'] == 'paid' and ex['pay_amount'] is not None:
        row_fixes['pay_amount'] = ex['pay_amount']

    if row_fixes:
        fixes[db_id] = {'fixes': row_fixes, 'handle': db_r['handle'],
                        'platform': db_r['platform'], 'campaign': db_r['campaign_code'],
                        'excel_row': ex['excel_row'], 'match_type': match_type,
                        'excel_platform': ex['platform'], 'excel_pic': ex['pic']}

# ---- Report ----
print(f"=== FIXES FOUND: {len(fixes)} rows ===\n")

field_counts = defaultdict(int)
for v in fixes.values():
    for f in v['fixes']:
        field_counts[f] += 1

print("Summary by field:")
for f, cnt in sorted(field_counts.items(), key=lambda x: -x[1]):
    print(f"  {f}: {cnt} rows")
print()

print("Detail:")
for db_id, info in sorted(fixes.items()):
    fields_str = ', '.join(f"{k}={v}" for k, v in info['fixes'].items())
    print(f"  id={db_id:4d} [{info['handle']}/{info['platform'] or 'NULL'}/camp={info['campaign']}]"
          f" (excel row {info['excel_row']}, {info['match_type']}): {fields_str}")

if warnings:
    print(f"\n=== WARNINGS ({len(warnings)}) ===")
    for w in warnings:
        print(f"  {w}")

# ---- Show posted rows with NULL post_url that weren't fixed (Excel also NULL) ----
posted_null_url = [r for r in db_rows if r['post_url'] is None and r['status'] == 'posted'
                   and r['id'] not in fixes]
if posted_null_url:
    print(f"\n=== POSTED rows with NULL post_url — Excel ก็ว่างจริง ({len(posted_null_url)} rows) ===")
    for r in posted_null_url:
        print(f"  id={r['id']:4d} [{r['handle']}/{r['platform']}/camp={r['campaign_code']}]")

# ---- Generate SQL ----
print("\n\n=== SQL FIXES ===")
sql_statements = []
for db_id, info in sorted(fixes.items()):
    fmap = info['fixes']
    parts = []
    for field, value in fmap.items():
        if value is None:
            parts.append(f"{field} = NULL")
        elif isinstance(value, str):
            escaped = value.replace("'", "''")
            parts.append(f"{field} = '{escaped}'")
        elif isinstance(value, Decimal):
            parts.append(f"{field} = {value}")
        else:
            parts.append(f"{field} = {value}")
    sql = f"UPDATE placements SET {', '.join(parts)} WHERE id = {db_id};"
    sql_statements.append(sql)

for s in sql_statements:
    print(s)

print(f"\n-- Total: {len(sql_statements)} UPDATE statements")

sql_path = r'D:\internship\KOL_management\kol-system\scripts\full_merge_fixes.sql'
with open(sql_path, 'w', encoding='utf-8') as f:
    f.write("-- Comprehensive merge-cell audit fixes\n")
    f.write("-- Generated by full_merge_audit.py\n\n")
    for s in sql_statements:
        f.write(s + "\n")
print(f"SQL written to: {sql_path}")

cur.close()
conn.close()
