"""
ดึง hyperlink จาก column F (KOL handle) ใน Excel
แล้วอัปเดต profile_url ใน kols table

วิธีใช้:
  python scripts/import_kol_profile_urls.py          # dry-run (แสดงผล ไม่ commit)
  python scripts/import_kol_profile_urls.py --apply  # apply จริง
"""
import sys, io, argparse
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
import psycopg2

EXCEL_PATH = r'C:\Users\USER-SHD-046\Downloads\Copy of 2026 Dreame Plan&Summary Update.xlsx'
DB_URL = "postgresql://postgres.hdrweioqqqpslsjizkci:Shd2025%21ffofo@aws-1-ap-northeast-1.pooler.supabase.com:5432/postgres"

parser = argparse.ArgumentParser()
parser.add_argument('--apply', action='store_true', help='commit ลง DB จริง')
args = parser.parse_args()

# Load Excel สองรอบ: data_only=True สำหรับค่า cell, data_only=False สำหรับ hyperlink
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws = wb['KOL Performance 2026']

wb_hl = openpyxl.load_workbook(EXCEL_PATH, data_only=False)
ws_hl = wb_hl['KOL Performance 2026']

# Build merge maps
merge_map = {}        # (row, col) -> top-left cell value
merge_top_left = {}   # (row, col) -> (tl_row, tl_col)
for mr in ws.merged_cells.ranges:
    tl_row, tl_col = mr.min_row, mr.min_col
    for row in range(mr.min_row, mr.max_row + 1):
        for col in range(mr.min_col, mr.max_col + 1):
            merge_map[(row, col)] = ws.cell(tl_row, tl_col).value
            merge_top_left[(row, col)] = (tl_row, tl_col)

def gv(row, col):
    return merge_map.get((row, col), ws.cell(row, col).value)

def get_hyperlink_col6(row):
    """ดึง hyperlink target จาก column F (col 6) — อ่านจาก top-left ของ merge range"""
    actual_row, actual_col = merge_top_left.get((row, 6), (row, 6))
    cell = ws_hl.cell(actual_row, actual_col)
    return cell.hyperlink.target if cell.hyperlink else None

def is_url(val):
    if not val: return False
    s = str(val).strip()
    return s.startswith('http://') or s.startswith('https://')

def norm(h):
    """normalize handle เหมือน server: lowercase + strip spaces (ไม่ตัด @)"""
    return str(h or '').strip().lower().replace(' ', '')

# ---- สแกน Excel: เก็บ handle_norm -> profile_url (first occurrence) ----
handle_url_map = {}   # handle_norm -> url
handle_first_row = {} # handle_norm -> excel row แรก

for r in range(3, ws.max_row + 1):
    kol = gv(r, 6)
    platform = gv(r, 3)
    if not kol or not platform:
        continue
    kol_str = str(kol).strip()
    if kol_str == str(platform).strip():
        continue
    low = kol_str.lower()
    if low.startswith('total') or low.startswith('รวม') or low.startswith('sum'):
        continue

    key = norm(kol_str)
    if key in handle_url_map:
        continue  # เอา first occurrence พอ

    url = get_hyperlink_col6(r)
    handle_url_map[key] = str(url).strip() if is_url(url) else None
    handle_first_row[key] = r

has_url_count = sum(1 for v in handle_url_map.values() if v)
print(f"KOL handles ใน Excel:  {len(handle_url_map)}")
print(f"มี profile URL:         {has_url_count}")
print(f"ไม่มี URL:              {len(handle_url_map) - has_url_count}")

if has_url_count:
    print("\nตัวอย่าง URL (10 รายการแรก):")
    shown = 0
    for k, v in handle_url_map.items():
        if v:
            print(f"  [{k[:30]:<30}] → {v[:70]}")
            shown += 1
            if shown >= 10:
                break

# ---- Match กับ DB ----
conn = psycopg2.connect(DB_URL)
cur = conn.cursor()
cur.execute("SELECT id, handle, handle_normalized, profile_url FROM kols ORDER BY id")
db_kols = cur.fetchall()

updates = []    # (kol_id, handle, new_url)
already_ok = 0
no_excel_url = 0
not_in_excel = 0

for kol_id, handle, handle_normalized, current_url in db_kols:
    # ลอง match ด้วย handle_normalized ก่อน แล้วค่อย handle (raw normalized)
    url = handle_url_map.get(handle_normalized or '') \
       or handle_url_map.get(norm(handle or ''))

    if url is None:
        # ไม่พบใน Excel หรือพบแต่ไม่มี URL
        key = handle_normalized or norm(handle or '')
        if key in handle_url_map:
            no_excel_url += 1   # Excel มีแต่ไม่มี hyperlink
        else:
            not_in_excel += 1   # ไม่พบใน Excel เลย (KOL เพิ่มเอง)
        continue

    if (current_url or '').strip() == url.strip():
        already_ok += 1
        continue

    updates.append((kol_id, handle, url))

print(f"\n=== ผลการ match ===")
print(f"  จะอัปเดต:              {len(updates)}")
print(f"  ตรงกันแล้ว:            {already_ok}")
print(f"  Excel ไม่มี hyperlink: {no_excel_url}")
print(f"  ไม่พบใน Excel:         {not_in_excel} (KOL ที่เพิ่มเอง)")

if updates:
    print(f"\nรายการที่จะอัปเดต:")
    for kol_id, handle, url in updates:
        print(f"  id={kol_id:4d} [{handle[:35]:<35}] → {url[:65]}")

if not args.apply:
    print(f"\n[DRY-RUN] ไม่ได้ commit — รัน --apply เพื่อบันทึกจริง")
    cur.close()
    conn.close()
    sys.exit(0)

# ---- Apply ----
if updates:
    for kol_id, _, url in updates:
        cur.execute("UPDATE kols SET profile_url = %s WHERE id = %s", (url, kol_id))
    conn.commit()
    print(f"\n✓ อัปเดต {len(updates)} rows เรียบร้อย")
else:
    print("\n✓ ไม่มีอะไรต้องอัปเดต")

cur.close()
conn.close()
