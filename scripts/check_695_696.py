"""
Check placements 695, 696: why are person_in_charge, platform_id, post_url null?
Search Excel by handle only (since platform_id is NULL in DB, can't use it for matching).
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
import psycopg2

EXCEL_PATH = r'C:\Users\USER-SHD-046\Downloads\Copy of 2026 Dreame Plan&Summary Update.xlsx'
DB_URL = "postgresql://postgres.hdrweioqqqpslsjizkci:Shd2025%21ffofo@aws-1-ap-northeast-1.pooler.supabase.com:5432/postgres"

def normalize_campaign(val):
    if val is None: return ''
    try:
        f = float(val)
        s = f'{round(f, 2):.2f}'.rstrip('0').rstrip('.')
        return s
    except:
        return str(val).strip()

# ---- Excel load (merge-cell aware) ----
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws = wb['KOL Performance 2026']
merge_map = {}
for mr in ws.merged_cells.ranges:
    tl = ws.cell(mr.min_row, mr.min_col)
    for row in range(mr.min_row, mr.max_row + 1):
        for col in range(mr.min_col, mr.max_col + 1):
            merge_map[(row, col)] = tl.value

def gv(r, c):
    return merge_map.get((r, c), ws.cell(r, c).value)

# Print first 5 rows to understand structure
print("First 5 rows of Excel (columns A-O):")
for r in range(1, 6):
    vals = [gv(r, c) for c in range(1, 16)]
    print(f"  row {r}: {vals}")
print()

# ---- DB ----
conn = psycopg2.connect(DB_URL)
cur = conn.cursor()
cur.execute("""
    SELECT p.id, p.kol_id, k.handle, p.platform_id, pl.name as platform_name,
           p.campaign_id, c.code as campaign_code,
           p.person_in_charge_id, p.person_in_charge,
           p.post_url, p.status, p.payment_type, p.placement_type,
           p.follower_at_time, p.final_price
    FROM placements p
    LEFT JOIN kols k ON k.id = p.kol_id
    LEFT JOIN platforms pl ON pl.id = p.platform_id
    LEFT JOIN campaigns c ON c.id = p.campaign_id
    WHERE p.id IN (695, 696)
    ORDER BY p.id
""")
rows = cur.fetchall()
col_names = [d[0] for d in cur.description]

print("=" * 70)
print("DB DATA FOR PLACEMENTS 695, 696")
print("=" * 70)
for row in rows:
    print()
    for col, val in zip(col_names, row):
        print(f"  {col:30s}: {val!r}")

# ---- Excel search by handle + campaign ----
target_handle = 'เจอนี่เจอนั่น-Journeyjournal'
target_campaign = '5.5'

print()
print(f"Searching Excel for handle='{target_handle}', campaign='{target_campaign}'")
print("(also doing loose search by handle partial match)")
print()

found = []
max_row = ws.max_row
for r in range(2, max_row + 1):
    handle_raw = gv(r, 6)
    if handle_raw is None:
        continue
    handle_str = str(handle_raw).strip()
    campaign_norm = normalize_campaign(gv(r, 1))

    # exact match
    if handle_str == target_handle and campaign_norm == target_campaign:
        found.append(('EXACT', r))
    # loose: handle contains key Thai text
    elif 'เจอนี่' in handle_str or 'Journeyjournal' in handle_str:
        found.append(('LOOSE', r))

print(f"Found {len(found)} match(es):")
for match_type, r in found:
    print(f"\n  [{match_type}] Excel row {r}:")
    for c, label in [(1,'A Campaign'),(2,'B PIC'),(3,'C Platform'),(4,'D Category'),
                     (5,'E Model'),(6,'F Handle'),(7,'G Tier'),(8,'H Follower'),
                     (9,'I Price'),(15,'O Post URL')]:
        merged_val = gv(r, c)
        raw_val = ws.cell(r, c).value
        print(f"    col {label:15s}: merged={merged_val!r}  raw={raw_val!r}")

# Show context around found rows
if found:
    print()
    print("=" * 70)
    print("CONTEXT (rows ±3 around matches)")
    print("=" * 70)
    shown = set()
    for _, er in found:
        for offset in range(-3, 4):
            rr = er + offset
            if rr < 2 or rr > max_row or rr in shown:
                continue
            shown.add(rr)
            marker = " <-- MATCH" if rr == er else ""
            handles_nearby = [gv(rr, 6)]
            print(f"  row {rr:4d}{marker}: "
                  f"campaign={gv(rr,1)!r} pic={gv(rr,2)!r} "
                  f"platform={gv(rr,3)!r} handle={gv(rr,6)!r} "
                  f"post_url={gv(rr,15)!r}")

conn.close()
print("\nDone.")
