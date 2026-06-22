"""
Sequential matching with fixed campaign code comparison (floating point).
Find all DB rows with NULL follower/final_price that can be filled from Excel.
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
import psycopg2
import re
from decimal import Decimal, InvalidOperation
from datetime import date, datetime

EXCEL_PATH = r'C:\Users\USER-SHD-046\Downloads\Copy of 2026 Dreame Plan&Summary Update.xlsx'
DB_URL = "postgresql://postgres.hdrweioqqqpslsjizkci:Shd2025%21ffofo@aws-1-ap-northeast-1.pooler.supabase.com:5432/postgres"

def normalize_campaign(val):
    """Convert 2.20000000000001 or '2.2' -> '2.2'"""
    if val is None: return ''
    try:
        f = float(val)
        # Round to 2 decimal places and strip trailing zeros
        s = f'{round(f, 2):.2f}'.rstrip('0').rstrip('.')
        # Handle X.0 case -> X (like 1.0 -> 1, but 2.2 -> 2.2)
        return s
    except: return str(val).strip()

def parse_follower(val):
    if val is None: return None
    s = str(val).strip().upper().replace(',', '')
    if s in ('', 'NONE', '-', 'N/A', '0', '0.0'): return None
    try:
        m = re.match(r'^([\d.]+)\s*([KMB]?)$', s)
        if m:
            num = float(m.group(1))
            mult = {'K': 1000, 'M': 1_000_000, 'B': 1_000_000_000, '': 1}[m.group(2)]
            v = int(num * mult)
            return v if v > 0 else None
    except: pass
    return None

def parse_price(val):
    if val is None: return None
    s = str(val).strip().upper()
    if s in ('FREE', 'BARTER', '-', '', 'N/A', 'NONE', '0', '0.0'): return None
    try:
        d = Decimal(s.replace(',', ''))
        return d if d > 0 else None
    except InvalidOperation: return None

def infer_payment_type(val):
    if val is None: return 'paid'
    s = str(val).strip().upper()
    if s == 'FREE': return 'free'
    if s == 'BARTER': return 'barter'
    return 'paid'

# ---- Excel load ----
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws = wb['KOL Performance 2026']
merge_map = {}
for mr in ws.merged_cells.ranges:
    tl = ws.cell(mr.min_row, mr.min_col)
    for row in range(mr.min_row, mr.max_row + 1):
        for col in range(mr.min_col, mr.max_col + 1):
            merge_map[(row, col)] = tl.value
def gv(row, col):
    return merge_map.get((row, col), ws.cell(row, col).value)

excel_rows = []
for r in range(3, ws.max_row + 1):
    kol = gv(r, 6)
    platform = gv(r, 3)
    if not kol or not platform: continue
    kol_str = str(kol).strip()
    plat_str = str(platform).strip()
    if kol_str == plat_str: continue
    low = kol_str.lower()
    if low.startswith('total') or low.startswith('รวม') or low.startswith('sum'): continue
    fp_raw = gv(r, 11)
    c_norm = normalize_campaign(gv(r, 1))
    excel_rows.append({
        'excel_row': r,
        'campaign_raw': gv(r, 1),
        'campaign': c_norm,
        'pic': str(gv(r, 2) or '').strip(),
        'platform': plat_str,
        'model': gv(r, 5),
        'kol_handle': kol_str,
        'follower_raw': gv(r, 9),
        'follower_int': parse_follower(gv(r, 9)),
        'final_price_raw': fp_raw,
        'final_price': parse_price(fp_raw),
        'payment_type': infer_payment_type(fp_raw),
        'pay_amount': parse_price(gv(r, 12)),
        'pay_amount_raw': gv(r, 12),
    })

print(f"Excel data rows: {len(excel_rows)}")

# ---- DB ----
conn = psycopg2.connect(DB_URL)
cur = conn.cursor()
cur.execute("""
    SELECT p.id, k.handle, plt.name, c.code,
           p.follower_at_time, p.final_price, p.pay_amount,
           p.payment_type, p.placement_type, p.status, p.campaign_id
    FROM placements p
    JOIN kols k ON k.id = p.kol_id
    JOIN platforms plt ON plt.id = p.platform_id
    LEFT JOIN campaigns c ON c.id = p.campaign_id
    ORDER BY p.id
""")
db_data = [{'id': r[0], 'handle': r[1], 'platform': r[2], 'campaign_code': r[3] or '',
             'follower_at_time': r[4], 'final_price': r[5], 'pay_amount': r[6],
             'payment_type': r[7], 'placement_type': r[8], 'status': r[9]}
           for r in cur.fetchall()]
print(f"DB placements: {len(db_data)}")

# ---- Match by (handle, platform, campaign) ----
from collections import defaultdict
excel_lookup = defaultdict(list)
for i, er in enumerate(excel_rows):
    key = (er['kol_handle'].lower(), er['platform'].lower(), er['campaign'])
    excel_lookup[key].append((i, er))

matched_pairs = []
unmatched_db = []

for db_r in db_data:
    c_code = db_r['campaign_code']
    key = (db_r['handle'].lower(), db_r['platform'].lower(), c_code)
    matches = excel_lookup.get(key, [])
    if matches:
        # If multiple matches (same KOL+platform+campaign), match by position in order
        matched_pairs.append((db_r, matches[0][1]))
    else:
        unmatched_db.append(db_r)

print(f"\nMatched: {len(matched_pairs)}/{len(db_data)}")
print(f"Unmatched: {len(unmatched_db)}")

# ---- Check unmatched ----
no_excel_match = []
wrong_campaign_only = []
for r in unmatched_db:
    key2 = (r['handle'].lower(), r['platform'].lower())
    excel_candidates = []
    for er in excel_rows:
        if er['kol_handle'].lower() == r['handle'].lower() and er['platform'].lower() == r['platform'].lower():
            excel_candidates.append(er['campaign'])
    if excel_candidates:
        wrong_campaign_only.append((r, excel_candidates))
    else:
        no_excel_match.append(r)

print(f"\nUnmatched - handle+platform exists in Excel (campaign mismatch): {len(wrong_campaign_only)}")
print(f"Unmatched - NO match in Excel at all: {len(no_excel_match)}")

if no_excel_match:
    print("\nNo-match DB rows (handle not in Excel):")
    for r in no_excel_match[:20]:
        print(f"  id={r['id']:4d} {r['handle']:<30} {r['platform']:<12} campaign={r['campaign_code']}")

# ---- Fixable follower NULLs from matched pairs ----
fol_fixes = []
fol_ambiguous = []
for db_r, ex_r in matched_pairs:
    if db_r['follower_at_time'] is None and ex_r['follower_int'] is not None:
        fol_fixes.append((db_r, ex_r))

# Also check unmatched rows - try by handle+platform+close campaign
for db_r in unmatched_db:
    if db_r['follower_at_time'] is not None: continue
    # Try handle+platform, pick excel rows with same campaign or NULL
    candidates = [er for er in excel_rows
                  if er['kol_handle'].lower() == db_r['handle'].lower()
                  and er['platform'].lower() == db_r['platform'].lower()
                  and er['follower_int'] is not None]
    unique_vals = list({e['follower_int'] for e in candidates})
    if len(unique_vals) == 1:
        fol_fixes.append((db_r, candidates[0]))
    elif len(unique_vals) > 1:
        fol_ambiguous.append((db_r, unique_vals))

print(f"\n=== follower_at_time fixes available: {len(fol_fixes)} ===")
print(f"    ambiguous (multiple values): {len(fol_ambiguous)}")
for db_r, ex_r in fol_fixes:
    fv = ex_r['follower_int'] if isinstance(ex_r, dict) else ex_r
    print(f"  id={db_r['id']:4d} {db_r['handle']:<30} {db_r['platform']:<12} -> {fv}")

if fol_ambiguous:
    print("\n  Ambiguous follower:")
    for db_r, vals in fol_ambiguous:
        print(f"  id={db_r['id']:4d} {db_r['handle']:<30} vals={vals}")

# ---- Fixable final_price NULLs ----
fp_fixes = []
fp_ambiguous = []
for db_r, ex_r in matched_pairs:
    if db_r['final_price'] is None and db_r['payment_type'] == 'paid' and ex_r['final_price'] is not None:
        fp_fixes.append((db_r, ex_r))

for db_r in unmatched_db:
    if db_r['final_price'] is not None or db_r['payment_type'] != 'paid': continue
    candidates = [er for er in excel_rows
                  if er['kol_handle'].lower() == db_r['handle'].lower()
                  and er['platform'].lower() == db_r['platform'].lower()
                  and er['final_price'] is not None]
    unique_vals = list({str(e['final_price']) for e in candidates})
    if len(unique_vals) == 1:
        fp_fixes.append((db_r, candidates[0]))
    elif len(unique_vals) > 1:
        fp_ambiguous.append((db_r, unique_vals))

print(f"\n=== final_price fixes available (paid, NULL): {len(fp_fixes)} ===")
print(f"    ambiguous: {len(fp_ambiguous)}")
for db_r, ex_r in fp_fixes:
    fv = ex_r['final_price'] if isinstance(ex_r, dict) else ex_r
    print(f"  id={db_r['id']:4d} {db_r['handle']:<30} {db_r['platform']:<12} -> {fv}")

if fp_ambiguous:
    print("\n  Ambiguous final_price:")
    for db_r, vals in fp_ambiguous:
        print(f"  id={db_r['id']:4d} {db_r['handle']:<30} vals={vals}")

# ---- Summary of remaining NULLs with no Excel data ----
still_null_fol = sum(1 for r in db_data if r['follower_at_time'] is None)
fixable_ids = {db_r['id'] for db_r, _ in fol_fixes}
truly_missing_fol = [r for r in db_data if r['follower_at_time'] is None and r['id'] not in fixable_ids]
print(f"\n=== follower_at_time: {still_null_fol} NULL total ===")
print(f"  Fixable from Excel: {len(fol_fixes)}")
print(f"  Truly missing (no Excel value): {len(truly_missing_fol)}")
if truly_missing_fol:
    print("  IDs with no Excel data:")
    for r in truly_missing_fol:
        print(f"    id={r['id']:4d} {r['handle']:<30} {r['platform']}")

cur.close()
conn.close()
