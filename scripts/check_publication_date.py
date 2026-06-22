"""
ตรวจ publication_date: เปรียบเทียบ cell value ใน col O (Excel) กับ DB
พร้อมแก้ปัญหา DD/MM vs MM/DD — Excel locale อเมริกัน interpret ผิด
แนวทาง: ถ้า day ≤ 12 (ambiguous) → swap month/day แล้วเช็ค campaign period
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
import psycopg2
from datetime import datetime, date
from collections import defaultdict

EXCEL_PATH = r'C:\Users\USER-SHD-046\Downloads\Copy of 2026 Dreame Plan&Summary Update.xlsx'
DB_URL = "postgresql://postgres.hdrweioqqqpslsjizkci:Shd2025%21ffofo@aws-1-ap-northeast-1.pooler.supabase.com:5432/postgres"

# Campaign code → approx month (สำหรับเช็ค plausibility)
CAMPAIGN_MONTH = {
    '1.1': 1, '1.15': 1, '1.25': 1,
    '2.2': 2, '2.15': 2, '2.25': 2,
    '3.3': 3, '3.15': 3, '3.25': 3,
    '4.4': 4, '4.15': 4, '4.25': 4,
    '5.5': 5, '5.15': 5, '5.25': 5,
    '6.6': 6, '6.15': 6, '6.25': 6,
    '7.7': 7, '7.15': 7, '7.25': 7,
    '8.8': 8, '8.15': 8, '8.25': 8,
    '9.9': 9, '9.15': 9, '9.25': 9,
    '10.10': 10, '10.15': 10, '10.25': 10,
    '11.11': 11, '11.15': 11, '11.25': 11,
    '12.12': 12, '12.15': 12, '12.25': 12,
}

def normalize_campaign(val):
    if val is None: return ''
    try:
        f = float(val)
        s = f'{round(f, 2):.2f}'.rstrip('0').rstrip('.')
        return s
    except:
        return str(val).strip()

def is_url(val):
    if not val: return False
    return str(val).strip().startswith('http')

def fix_date(d, campaign_code):
    """
    รับ date จาก openpyxl (อาจ swap MM/DD) คืนวันที่ที่ถูกต้อง
    ถ้า day ≤ 12 (ambiguous) → ลอง swap แล้วเช็คว่า month ตรง campaign
    """
    if d is None: return None, 'none'
    if d.year < 2020 or d.year > 2030: return None, 'bad_year'

    camp_month = CAMPAIGN_MONTH.get(campaign_code)

    # วันชัดเจน (day > 12) — Excel interpret ถูกแน่
    if d.day > 12:
        return d, 'unambiguous'

    # วันกำกวม (day ≤ 12, month ≤ 12) — ลอง swap
    try:
        swapped = date(d.year, d.day, d.month)  # swap month↔day
    except ValueError:
        return d, 'swap_invalid'

    if camp_month is None:
        # ไม่รู้ campaign month — เลือก swapped (Thai intent)
        return swapped, 'swapped_no_camp'

    # เลือกอันที่ month ตรงกับ campaign (ยอมให้ผิด ±1 เดือน)
    orig_ok   = abs(d.month - camp_month) <= 1
    swap_ok   = abs(swapped.month - camp_month) <= 1

    if swap_ok and not orig_ok:
        return swapped, 'swapped'
    elif orig_ok and not swap_ok:
        return d, 'original_ok'
    elif swap_ok and orig_ok:
        # ทั้งคู่เป็นไปได้ — ใช้ swapped (Thai intent)
        return swapped, 'both_ok_use_swap'
    else:
        # ไม่มีอันไหนตรง — ใช้ swapped (Thai intent)
        return swapped, 'neither_ok_use_swap'

# ---- Load Excel ----
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws = wb['KOL Performance 2026']

merge_map = {}
merge_top_left = {}
for mr in ws.merged_cells.ranges:
    tl_row, tl_col = mr.min_row, mr.min_col
    tl_val = ws.cell(tl_row, tl_col).value
    for row in range(mr.min_row, mr.max_row + 1):
        for col in range(mr.min_col, mr.max_col + 1):
            merge_map[(row, col)] = tl_val
            merge_top_left[(row, col)] = (tl_row, tl_col)

def gv(row, col):
    return merge_map.get((row, col), ws.cell(row, col).value)

def get_pub_date_raw(row):
    actual_row, _ = merge_top_left.get((row, 15), (row, 15))
    raw = ws.cell(actual_row, 15).value
    if is_url(raw): return None
    if isinstance(raw, datetime): return raw.date()
    if isinstance(raw, date): return raw
    # parse text fallbacks
    for fmt in ('%d/%m/%Y', '%d/%m/%y', '%Y-%m-%d'):
        try:
            d = datetime.strptime(str(raw).strip(), fmt).date()
            if 2020 <= d.year <= 2030:
                return d
        except (ValueError, TypeError):
            pass
    return None

# ---- Read Excel rows ----
excel_rows = []
for r in range(3, ws.max_row + 1):
    kol = gv(r, 6)
    platform = gv(r, 3)
    if not kol or not platform: continue
    kol_str = str(kol).strip()
    plat_str = str(platform).strip()
    if kol_str == plat_str: continue
    low = kol_str.lower()
    if low.startswith('total') or low.startswith('รวม') or low.startswith('sum'): continue
    camp = normalize_campaign(gv(r, 1))
    raw_date = get_pub_date_raw(r)
    excel_rows.append({
        'excel_row': r,
        'campaign': camp,
        'platform': plat_str,
        'handle': kol_str,
        'raw_date': raw_date,
    })

# ---- DB ----
conn = psycopg2.connect(DB_URL)
cur = conn.cursor()
cur.execute("""
    SELECT p.id, k.handle, plt.name, c.code, p.publication_date, p.status
    FROM placements p
    JOIN kols k ON k.id = p.kol_id
    LEFT JOIN platforms plt ON plt.id = p.platform_id
    LEFT JOIN campaigns c ON c.id = p.campaign_id
    ORDER BY p.id
""")
db_by_3key = defaultdict(list)
for r in cur.fetchall():
    key = (r[1].lower(), (r[2] or '').lower(), r[3] or '')
    db_by_3key[key].append({
        'id': r[0], 'handle': r[1], 'platform': r[2] or '',
        'campaign': r[3] or '', 'publication_date': r[4], 'status': r[5]
    })

# ---- Compare & fix ----
excel_used = set()
fixes = []
mismatch_swap = []
already_ok = 0
reason_counts = defaultdict(int)

for ep in excel_rows:
    k3 = (ep['handle'].lower(), ep['platform'].lower(), ep['campaign'])
    candidates = [d for d in db_by_3key.get(k3, []) if d['id'] not in excel_used]
    if not candidates or ep['raw_date'] is None: continue
    dr = candidates[0]
    excel_used.add(dr['id'])

    corrected, reason = fix_date(ep['raw_date'], ep['campaign'])
    reason_counts[reason] += 1

    if corrected is None: continue

    db_date = dr['publication_date']
    if db_date is None:
        fixes.append({**dr, 'new_date': corrected, 'reason': reason, 'excel_row': ep['excel_row']})
    elif db_date == corrected:
        already_ok += 1
    else:
        mismatch_swap.append({**dr, 'ex_corrected': corrected, 'ex_raw': ep['raw_date'], 'reason': reason})

print(f"=== สรุปผล publication_date ===")
print(f"  ตรงกัน (DB = Excel corrected): {already_ok}")
print(f"  ต้องเพิ่ม (DB NULL):            {len(fixes)}")
print(f"  ยังต่างกันหลัง correct:         {len(mismatch_swap)}")
print(f"\nReason breakdown:")
for r, c in sorted(reason_counts.items(), key=lambda x: -x[1]):
    print(f"  {r}: {c}")

if mismatch_swap:
    print(f"\nตัวอย่าง mismatch ที่เหลือ (หลัง swap fix, 10 แถว):")
    for m in mismatch_swap[:10]:
        print(f"  id={m['id']:4d} [{m['handle'][:20]}/camp={m['campaign']}]: DB={m['publication_date']}  Excel_raw={m['ex_raw']}  corrected={m['ex_corrected']}  ({m['reason']})")

print(f"\nตัวอย่าง fixes (10 แถวแรก):")
for f in fixes[:10]:
    print(f"  id={f['id']:4d} [{f['handle'][:25]}/camp={f['campaign']}]: NULL → {f['new_date']}  ({f['reason']})")

if fixes:
    sql_path = r'D:\internship\KOL_management\kol-system\scripts\publication_date_fixes.sql'
    with open(sql_path, 'w', encoding='utf-8') as fout:
        fout.write("-- publication_date fixes (DD/MM-aware)\n\n")
        for f in fixes:
            fout.write(f"UPDATE placements SET publication_date = '{f['new_date']}' WHERE id = {f['id']};\n")
    print(f"\nSQL written to: {sql_path}  ({len(fixes)} statements)")

cur.close()
conn.close()
