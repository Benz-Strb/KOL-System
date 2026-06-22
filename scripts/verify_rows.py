import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\USER-SHD-046\Downloads\Copy of 2026 Dreame Plan&Summary Update.xlsx', data_only=True)
ws = wb['KOL Performance 2026']
merge_map = {}
for mr in ws.merged_cells.ranges:
    tl = ws.cell(mr.min_row, mr.min_col)
    for row in range(mr.min_row, mr.max_row + 1):
        for col in range(mr.min_col, mr.max_col + 1):
            merge_map[(row, col)] = tl.value
def gv(r, c): return merge_map.get((r, c), ws.cell(r, c).value)
def norm(v):
    if v is None: return ''
    try: return str(round(float(v), 2)).rstrip('0').rstrip('.')
    except: return str(v).strip()

rows_to_check = [220, 240, 342, 349, 454, 594, 735, 816, 1031]
print('Verifying follower fix source rows:')
for r in rows_to_check:
    kol = str(gv(r, 6) or '')
    plat = str(gv(r, 3) or '')
    print(f'  row {r:4d}: campaign={norm(gv(r,1)):<6} platform={plat:<12} kol={kol:<35} follower={gv(r,9)}')
