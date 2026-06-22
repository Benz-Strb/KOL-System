"""
Full field-by-field verification: DB vs Excel (merge-cell aware).
For every matched row, compare platform, PIC, follower, final_price,
payment_type, pay_amount, and post_url (posted rows only).
Reports actual mismatches — not just NULLs.
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
import psycopg2
import re
from decimal import Decimal, InvalidOperation
from collections import defaultdict

EXCEL_PATH = r'C:\Users\USER-SHD-046\Downloads\Copy of 2026 Dreame Plan&Summary Update.xlsx'
DB_URL = "postgresql://postgres.hdrweioqqqpslsjizkci:Shd2025%21ffofo@aws-1-ap-northeast-1.pooler.supabase.com:5432/postgres"

# ---- helpers ----
def normalize_campaign(val):
    if val is None: return ''
    try:
        f = float(val)
        s = f'{round(f, 2):.2f}'.rstrip('0').rstrip('.')
        return s
    except:
        return str(val).strip()

def parse_follower(val):
    if val is None: return None
    s = str(val).strip().upper().replace(',', '')
    if s in ('', 'NONE', '-', 'N/A', '0', '0.0'): return None
    try:
        m = re.match(r'^([\d.]+)\s*([KMB]?)$', s)
        if m:
            num = float(m.group(1))
            mult = {'K': 1000, 'M': 1_000_000, 'B': 1_000_000_000, '': 1}[m.group(2)]
            v = int(num * mult)
            return v if v > 0 else None
    except:
        pass
    return None

def parse_price(val):
    if val is None: return None
    s = str(val).strip().upper()
    if s in ('FREE', 'BARTER', '-', '', 'N/A', 'NONE', '0', '0.0'): return None
    try:
        d = Decimal(str(s).replace(',', ''))
        return d if d > 0 else None
    except InvalidOperation:
        return None

def infer_payment_type(val):
    if val is None: return 'paid'
    s = str(val).strip().upper()
    if s == 'FREE': return 'free'
    if s == 'BARTER': return 'barter'
    return 'paid'

def is_url(val):
    if val is None: return False
    s = str(val).strip()
    return s.startswith('http://') or s.startswith('https://')

def dec_eq(a, b, tol=Decimal('0.01')):
    """Compare two Decimal/None values with tolerance."""
    if a is None and b is None: return True
    if a is None or b is None: return False
    try:
        return abs(Decimal(str(a)) - Decimal(str(b))) <= tol
    except:
        return False

# ---- Load Excel ----
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws = wb['KOL Performance 2026']

merge_map = {}
for mr in ws.merged_cells.ranges:
    tl = ws.cell(mr.min_row, mr.min_col)
    for row in range(mr.min_row, mr.max_row + 1):
        for col in range(mr.min_col, mr.max_col + 1):
            merge_map[(row, col)] = tl.value

def gv(row, col):
    return merge_map.get((row, col), ws.cell(row, col).value)

excel_rows = []
for r in range(3, ws.max_row + 1):
    kol = gv(r, 6)
    platform = gv(r, 3)
    if not kol or not platform:
        continue
    kol_str = str(kol).strip()
    plat_str = str(platform).strip()
    if kol_str == plat_str:
        continue
    low = kol_str.lower()
    if low.startswith('total') or low.startswith('รวม') or low.startswith('sum'):
        continue

    fp_raw = gv(r, 11)
    post_raw = ws.cell(r, 15).value
    post_url = str(post_raw).strip() if is_url(post_raw) else None

    excel_rows.append({
        'excel_row':    r,
        'campaign':     normalize_campaign(gv(r, 1)),
        'pic':          str(gv(r, 2) or '').strip(),
        'platform':     plat_str,
        'kol_handle':   kol_str,
        'follower_int': parse_follower(gv(r, 9)),
        'fp_raw':       str(fp_raw or '').strip().upper(),
        'final_price':  parse_price(fp_raw),
        'payment_type': infer_payment_type(fp_raw),
        'pay_amount':   parse_price(gv(r, 12)),
        'post_url':     post_url,
    })

# ---- DB ----
conn = psycopg2.connect(DB_URL)
cur = conn.cursor()

cur.execute("SELECT id, name FROM platforms")
plat_id_to_name = {r[0]: r[1] for r in cur.fetchall()}
plat_name_to_id = {v.lower(): k for k, v in plat_id_to_name.items()}

cur.execute("SELECT id, full_name FROM users")
user_id_to_name = {r[0]: r[1] for r in cur.fetchall()}
user_name_to_id = {v.lower(): k for k, v in user_id_to_name.items()}

cur.execute("""
    SELECT p.id, k.handle, plt.name, c.code,
           p.platform_id, p.person_in_charge_id, p.person_in_charge,
           p.follower_at_time, p.final_price, p.pay_amount, p.payment_type,
           p.post_url, p.placement_type, p.status
    FROM placements p
    JOIN kols k ON k.id = p.kol_id
    LEFT JOIN platforms plt ON plt.id = p.platform_id
    LEFT JOIN campaigns c ON c.id = p.campaign_id
    ORDER BY p.id
""")
db_rows = []
for r in cur.fetchall():
    db_rows.append({
        'id': r[0], 'handle': r[1], 'platform': r[2] or '', 'campaign_code': r[3] or '',
        'platform_id': r[4], 'person_in_charge_id': r[5], 'person_in_charge': r[6],
        'follower_at_time': r[7], 'final_price': r[8], 'pay_amount': r[9],
        'payment_type': r[10], 'post_url': r[11],
        'placement_type': r[12], 'status': r[13],
    })

print(f"Excel rows: {len(excel_rows)} | DB rows: {len(db_rows)}\n")

# ---- Build lookup ----
excel_by_3key = defaultdict(list)
excel_by_2key = defaultdict(list)
for er in excel_rows:
    k3 = (er['kol_handle'].lower(), er['platform'].lower(), er['campaign'])
    k2 = (er['kol_handle'].lower(), er['campaign'])
    excel_by_3key[k3].append(er)
    excel_by_2key[k2].append(er)

excel_used = set()

# ---- Compare ----
mismatches = []   # list of dicts
unmatched_db = []

for db_r in db_rows:
    if db_r['platform_id'] is None:
        k2 = (db_r['handle'].lower(), db_r['campaign_code'])
        candidates = [e for e in excel_by_2key.get(k2, []) if e['excel_row'] not in excel_used]
    else:
        k3 = (db_r['handle'].lower(), db_r['platform'].lower(), db_r['campaign_code'])
        candidates = [e for e in excel_by_3key.get(k3, []) if e['excel_row'] not in excel_used]

    if not candidates:
        unmatched_db.append(db_r)
        continue

    ex = candidates[0]
    excel_used.add(ex['excel_row'])

    issues = []

    # --- platform ---
    ex_plat_id = plat_name_to_id.get(ex['platform'].lower())
    if ex_plat_id and db_r['platform_id'] != ex_plat_id:
        issues.append({
            'field': 'platform',
            'db': db_r['platform'] or 'NULL',
            'excel': ex['platform'],
            'fix': f"platform_id = {ex_plat_id}",
        })

    # --- person_in_charge ---
    ex_pic_id = user_name_to_id.get(ex['pic'].lower()) if ex['pic'] else None
    db_pic_name = user_id_to_name.get(db_r['person_in_charge_id'], '')
    if ex['pic'] and ex_pic_id and db_r['person_in_charge_id'] != ex_pic_id:
        issues.append({
            'field': 'person_in_charge',
            'db': db_pic_name or 'NULL',
            'excel': ex['pic'],
            'fix': f"person_in_charge_id = {ex_pic_id}, person_in_charge = '{ex['pic']}'",
        })

    # --- follower_at_time ---
    if ex['follower_int'] is not None and db_r['follower_at_time'] != ex['follower_int']:
        issues.append({
            'field': 'follower_at_time',
            'db': db_r['follower_at_time'],
            'excel': ex['follower_int'],
            'fix': f"follower_at_time = {ex['follower_int']}",
        })

    # --- payment_type ---
    if db_r['payment_type'] != ex['payment_type']:
        issues.append({
            'field': 'payment_type',
            'db': db_r['payment_type'],
            'excel': ex['payment_type'],
            'fix': f"payment_type = '{ex['payment_type']}'",
        })

    # --- final_price (paid rows only) ---
    if db_r['payment_type'] == 'paid' and ex['payment_type'] == 'paid':
        if not dec_eq(db_r['final_price'], ex['final_price']):
            issues.append({
                'field': 'final_price',
                'db': db_r['final_price'],
                'excel': ex['final_price'],
                'fix': f"final_price = {ex['final_price']}" if ex['final_price'] else "final_price = NULL",
            })

    # --- pay_amount (paid rows only) ---
    if db_r['payment_type'] == 'paid' and ex['payment_type'] == 'paid':
        if not dec_eq(db_r['pay_amount'], ex['pay_amount']):
            # Only flag if Excel has a value (DB=NULL+Excel=NULL is fine)
            if ex['pay_amount'] is not None or db_r['pay_amount'] is not None:
                issues.append({
                    'field': 'pay_amount',
                    'db': db_r['pay_amount'],
                    'excel': ex['pay_amount'],
                    'fix': f"pay_amount = {ex['pay_amount']}" if ex['pay_amount'] else "pay_amount = NULL",
                })

    # --- post_url (posted rows only) ---
    if db_r['status'] == 'posted' and ex['post_url']:
        db_url = (db_r['post_url'] or '').strip()
        ex_url = ex['post_url'].strip()
        if db_url != ex_url:
            escaped = ex_url.replace("'", "''")
            issues.append({
                'field': 'post_url',
                'db': db_url or 'NULL',
                'excel': ex_url,
                'fix': f"post_url = '{escaped}'",
            })

    if issues:
        mismatches.append({
            'id': db_r['id'],
            'handle': db_r['handle'],
            'platform': db_r['platform'],
            'campaign': db_r['campaign_code'],
            'excel_row': ex['excel_row'],
            'issues': issues,
        })

# ---- Report ----
print(f"=== MISMATCHES: {len(mismatches)} rows ===\n")

field_counts = defaultdict(int)
for m in mismatches:
    for iss in m['issues']:
        field_counts[iss['field']] += 1

if field_counts:
    print("Summary by field:")
    for f, cnt in sorted(field_counts.items(), key=lambda x: -x[1]):
        print(f"  {f}: {cnt} rows")
    print()

for m in mismatches:
    ctx = f"id={m['id']:4d} [{m['handle']}/{m['platform']}/camp={m['campaign']}] (excel row {m['excel_row']})"
    for iss in m['issues']:
        print(f"  {ctx}")
        print(f"    {iss['field']}: DB={iss['db']}  EXCEL={iss['excel']}")
        print(f"    FIX: {iss['fix']}")

if unmatched_db:
    print(f"\n=== UNMATCHED DB ROWS (ไม่พบใน Excel): {len(unmatched_db)} ===")
    for r in unmatched_db:
        print(f"  id={r['id']:4d} [{r['handle']}/{r['platform']}/camp={r['campaign_code']}]")

# ---- Generate SQL ----
sql_statements = []
for m in mismatches:
    set_parts = []
    for iss in m['issues']:
        # Extract SET clause from fix string
        set_parts.append(iss['fix'])
    sql = f"UPDATE placements SET {', '.join(set_parts)} WHERE id = {m['id']};"
    sql_statements.append((m['id'], sql))

if sql_statements:
    print(f"\n\n=== SQL TO FIX ({len(sql_statements)} statements) ===")
    for _, s in sql_statements:
        print(s)

    sql_path = r'D:\internship\KOL_management\kol-system\scripts\verify_fixes.sql'
    with open(sql_path, 'w', encoding='utf-8') as f:
        f.write("-- Field-by-field verify fixes\n-- Generated by verify_all_fields.py\n\n")
        for _, s in sql_statements:
            f.write(s + "\n")
    print(f"\nSQL written to: {sql_path}")
else:
    print("\n✓ ไม่มี mismatch — DB ตรงกับ Excel ทุกแถวที่ match ได้")

print(f"\nSummary:")
print(f"  Matched:   {len(db_rows) - len(unmatched_db)}/{len(db_rows)}")
print(f"  Unmatched: {len(unmatched_db)}")
print(f"  Mismatch:  {len(mismatches)}")

cur.close()
conn.close()
